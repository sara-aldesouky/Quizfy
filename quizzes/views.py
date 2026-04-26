"""
=============================================================================
Quizfy Views
=============================================================================

This module contains all view functions for the Quizfy quiz platform.

View Categories:
----------------
PUBLIC VIEWS:
    - home()                    : Home page with role-based redirect
    - landing()                 : Marketing landing page

QUIZ TAKING VIEWS:
    - take_quiz()               : Student quiz-taking interface
    - quiz_join()               : QR code join page
    - quiz_scan_redirect()      : Handle QR scan authentication
    - quiz_result()             : Show quiz results to student
    - quiz_status()             : AJAX endpoint for quiz status
    - quiz_qr_code()            : Generate QR code image

TEACHER AUTHENTICATION:
    - teacher_signup()          : Register teacher account
    - teacher_login()           : Teacher login page
    - teacher_logout()          : Logout teacher

TEACHER QUIZ MANAGEMENT:
    - teacher_quizzes()         : Teacher dashboard (folders + quizzes)
    - create_quiz()             : Create new quiz
    - teacher_quiz_detail()     : View quiz details
    - delete_quiz()             : Delete a quiz
    - toggle_quiz_active()      : Start/stop quiz
    - edit_quiz_settings()      : Edit timer, due date, etc.
    - move_quiz()               : Move quiz to folder

TEACHER QUESTION MANAGEMENT:
    - create_question()         : Add question to quiz
    - edit_question()           : Edit existing question
    - delete_question()         : Delete question

TEACHER SUBMISSION MANAGEMENT:
    - quiz_submissions()        : List all submissions for a quiz
    - grade_submission()        : Grade student's submission
    - allow_extra_attempt()     : Give student extra attempt
    - adjust_attempts()         : Adjust attempt count (+/-)
    - delete_teacher_file()     : Remove teacher feedback file

TEACHER FOLDER MANAGEMENT:
    - create_folder()           : Create subject folder
    - folder_detail()           : View folder contents
    - folder_analytics()        : AI-powered analytics
    - delete_folder()           : Delete folder

TEACHER EXPORT FUNCTIONS:
    - export_submissions_excel(): Export quiz submissions
    - export_folder_boxes_excel(): Export folder boxes report
    - export_student_folder_excel(): Export student report

STUDENT VIEWS:
    - student_signup()          : Register student account
    - student_login()           : Student login page
    - student_dashboard()       : Student dashboard
    - student_submission_detail(): View submission details
    - enter_quiz()              : Enter quiz code form

PASSWORD MANAGEMENT:
    - change_password()         : Change password form
    - change_password_done()    : Password change success

UTILITY VIEWS:
    - teacher_help_bot()        : AI help assistant
    - email_diagnostic()        : Email config debug
    - send_test_email()         : Test email sending
    - app_logout()              : Universal logout

Author: Quizfy Team
=============================================================================
"""

# =============================================================================
# IMPORTS
# =============================================================================

# Django core imports
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
from django.contrib import messages
from django.urls import reverse
from django.http import HttpResponse, JsonResponse
from django.views.decorators.csrf import ensure_csrf_cookie, csrf_exempt
from django.views.decorators.http import require_POST
from django.db import transaction
from django.db.models import Count, Q
from django.utils import timezone
from django.conf import settings
from django.core.mail import send_mail

# Third-party imports
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
import qrcode
from io import BytesIO

# Python standard library
import json
import difflib
import re
import logging

# Local imports
from .models import (
    Quiz, Question, Submission, Answer, 
    StudentProfile, SubjectFolder, QuizAttemptPermission, FileSubmission,
    QuizSecurityViolation,
)
from .forms import (
    TeacherLoginForm, TeacherSignupForm, QuizForm, QuestionForm,
    EnterQuizForm, StudentSignupForm, StudentLoginForm, FolderForm,
    MoveQuizForm, QuizSettingsForm, ChangePasswordForm,
    TrueFalseQuestionForm, FileUploadSubmissionForm, FileUploadQuestionForm
)


# =============================================================================
# LOGGING SETUP
# =============================================================================

logger = logging.getLogger(__name__)

SECURE_QUIZ_AUTO_SUBMIT_THRESHOLD = 3
SECURE_QUIZ_SUSPICIOUS_THRESHOLD = 1
SECURE_QUIZ_HIGH_RISK_THRESHOLD = 3


# =============================================================================
# AUTHENTICATION DECORATORS
# =============================================================================

def student_required(view_func):
    """
    Decorator that restricts access to authenticated students only.
    Redirects to student login if not authenticated or not a student.
    """
    return user_passes_test(
        lambda u: u.is_authenticated and not u.is_staff and hasattr(u, "student_profile"),
        login_url="/student/login/"
    )(view_func)


def staff_required(view_func):
    """
    Decorator that restricts access to staff (teachers) only.
    Redirects to teacher login if not authenticated or not staff.
    """
    return user_passes_test(
        lambda u: u.is_authenticated and u.is_staff,
        login_url="/teacher/login/"
    )(view_func)


# =============================================================================
# PUBLIC VIEWS
# =============================================================================

def home(request):
    """
    Home page view with role-based redirects.
    
    - Authenticated teachers → Teacher dashboard
    - Authenticated students → Student dashboard  
    - Anonymous users → Landing page
    """
    if request.user.is_authenticated:
        if request.user.is_staff:
            return redirect("teacher_quizzes")
        return redirect("student_dashboard")
    return render(request, "quizzes/landing.html")


def landing(request):
    """
    Render the public landing page without redirecting authenticated users.
    This lets the logo/home link always go to the marketing landing page.
    """
    return render(request, "quizzes/landing.html")


# =============================================================================
# QR CODE GENERATION
# =============================================================================

@staff_required
def quiz_qr_code(request, quiz_code):
    """
    Generate and serve a QR code image for a quiz.
    
    The QR code contains a URL that students can scan to quickly
    access the quiz join page.
    """
    quiz = get_object_or_404(Quiz, code=quiz_code.upper(), teacher=request.user)
    
    try:
        # Build absolute URL for QR code (points to join page)
        protocol = "https" if request.is_secure() else "http"
        domain = request.get_host()
        qr_data = f"{protocol}://{domain}/quiz/{quiz.code}/join/"
        
        qr = qrcode.QRCode(version=1, box_size=10, border=2)
        qr.add_data(qr_data)
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white")
        
        # Serve as PNG
        buffer = BytesIO()
        img.save(buffer, format="PNG")
        buffer.seek(0)
        
        return HttpResponse(buffer.getvalue(), content_type="image/png")
    except Exception as e:
        # Return a blank 1x1 pixel if generation fails
        logger.error(f"QR code generation failed for {quiz_code}: {e}")
        return HttpResponse(
            bytes.fromhex("89504e470d0a1a0a0000000d494844520000000100000001080202000090773db30000000c49444154785e6300010000050001000b0b80c30000000049454e44ae426082"),
            content_type="image/png"
        )


# =============================================================================
# TEACHER HELP BOT (Smart FAQ System)
# =============================================================================

TEACHER_KB = [
    # ========== FOLDERS & ORGANIZATION ==========
    {
        "tags": ["create folder", "subject folder", "add subject", "new folder", "folder", "make folder"],
        "q": "How do I create a subject folder?",
        "a": (
            "📁 **Creating a Subject Folder:**\n\n"
            "1. Go to your **Dashboard** (My Subjects)\n"
            "2. Click **+ Create Subject Folder**\n"
            "3. Enter the folder name (e.g., Physics, Math 101)\n"
            "4. Click **Save**\n\n"
            "💡 Tip: Organize your quizzes by subject or course for easy management!"
        )
    },
    {
        "tags": ["move quiz", "assign quiz", "put quiz in folder", "organize quizzes", "ungrouped", "move to folder"],
        "q": "How do I move a quiz into a folder?",
        "a": (
            "📦 **Moving a Quiz to a Folder:**\n\n"
            "1. Find the quiz in **Ungrouped Quizzes** section\n"
            "2. Click the **Move** button\n"
            "3. Select the destination folder\n"
            "4. Click **Save**\n\n"
            "💡 You can also move quizzes between folders the same way!"
        )
    },
    {
        "tags": ["delete folder", "remove folder", "folder delete"],
        "q": "How do I delete a folder?",
        "a": (
            "🗑️ **Deleting a Folder:**\n\n"
            "1. Click the **🗑️** button on the folder card\n"
            "2. Choose what to do with quizzes inside:\n"
            "   • **Delete All**: Removes folder AND all quizzes\n"
            "   • **Move to Ungrouped**: Keeps quizzes, deletes only folder\n\n"
            "⚠️ Warning: Deleting quizzes will also delete all student submissions!"
        )
    },
    
    # ========== QUIZ CREATION ==========
    {
        "tags": ["create quiz", "new quiz", "make quiz", "add quiz"],
        "q": "How do I create a quiz?",
        "a": (
            "📝 **Creating a New Quiz:**\n\n"
            "1. Click **+ Create Quiz** (on dashboard or inside a folder)\n"
            "2. Enter the quiz **Title**\n"
            "3. Select **Quiz Type**:\n"
            "   • Multiple Choice\n"
            "   • True/False\n"
            "   • File Upload (students submit PDFs/images)\n"
            "4. Optionally assign to a folder\n"
            "5. Set **Duration** (in minutes) if timed\n"
            "6. Click **Create**\n\n"
            "💡 After creating, add questions from the quiz detail page!"
        )
    },
    {
        "tags": ["add question", "create question", "new question", "questions"],
        "q": "How do I add questions to a quiz?",
        "a": (
            "❓ **Adding Questions:**\n\n"
            "1. Open the quiz (click **Open**)\n"
            "2. Click **+ Add Question**\n"
            "3. Choose question type:\n"
            "   • **Multiple Choice**: 4 options, select correct one\n"
            "   • **True/False**: Just enter the statement\n"
            "   • **File Upload**: Students upload their answer as PDF/image\n"
            "4. Enter the question text\n"
            "5. Add an image (optional)\n"
            "6. Click **Save**\n\n"
            "💡 You can mix different question types in the same quiz!"
        )
    },
    {
        "tags": ["edit question", "change question", "modify question", "update question"],
        "q": "How do I edit a question?",
        "a": (
            "✏️ **Editing Questions:**\n\n"
            "1. Open the quiz\n"
            "2. Find the question you want to edit\n"
            "3. Click the **Edit** button\n"
            "4. Make your changes\n"
            "5. Click **Save**\n\n"
            "⚠️ Note: Editing after students have submitted may affect their results!"
        )
    },
    {
        "tags": ["delete question", "remove question"],
        "q": "How do I delete a question?",
        "a": (
            "🗑️ **Deleting Questions:**\n\n"
            "1. Open the quiz\n"
            "2. Find the question\n"
            "3. Click the **Delete** button\n"
            "4. Confirm deletion\n\n"
            "⚠️ Warning: This will also delete all student answers for that question!"
        )
    },
    {
        "tags": ["quiz type", "file upload quiz", "multiple choice quiz", "true false quiz", "quiz types", "type of quiz"],
        "q": "What quiz types are available?",
        "a": (
            "📋 **Available Quiz Types:**\n\n"
            "**1. Multiple Choice**\n"
            "   • 4 options per question\n"
            "   • Auto-graded instantly\n\n"
            "**2. True/False**\n"
            "   • Simple yes/no questions\n"
            "   • Auto-graded instantly\n\n"
            "**3. File Upload**\n"
            "   • Students upload PDF, JPG, or PNG\n"
            "   • You grade manually with comments\n"
            "   • Great for essays, drawings, handwritten work\n\n"
            "💡 You can also mix question types within a quiz!"
        )
    },
    
    # ========== QUIZ SETTINGS ==========
    {
        "tags": ["quiz settings", "timer", "duration", "time limit", "timed quiz", "minutes"],
        "q": "How do I set a time limit for a quiz?",
        "a": (
            "⏱️ **Setting a Time Limit:**\n\n"
            "1. Open the quiz\n"
            "2. Click **⚙️ Settings**\n"
            "3. Enter **Duration** in minutes\n"
            "4. Click **Save**\n\n"
            "How it works:\n"
            "• Timer starts when student opens the quiz\n"
            "• Auto-submits when time runs out\n"
            "• Students see countdown on their screen\n\n"
            "💡 Leave blank for unlimited time!"
        )
    },
    {
        "tags": ["due date", "deadline", "quiz expires", "end date", "close date"],
        "q": "How do I set a due date for a quiz?",
        "a": (
            "📅 **Setting a Due Date:**\n\n"
            "1. Open the quiz\n"
            "2. Click **⚙️ Settings**\n"
            "3. Set the **Due Date/Time**\n"
            "4. Click **Save**\n\n"
            "After the due date:\n"
            "• Quiz automatically closes\n"
            "• Students cannot submit anymore\n"
            "• Existing submissions are saved"
        )
    },
    {
        "tags": ["start quiz", "stop quiz", "activate", "deactivate", "toggle", "close quiz", "open quiz", "active"],
        "q": "How do I start or stop a quiz?",
        "a": (
            "🔄 **Starting/Stopping a Quiz:**\n\n"
            "1. Open the quiz\n"
            "2. Click the **Start Quiz** or **Stop Quiz** button\n\n"
            "**When Active (Started):**\n"
            "• Students can access and submit\n"
            "• Shows green indicator\n\n"
            "**When Inactive (Stopped):**\n"
            "• Students see 'Quiz Closed' message\n"
            "• No new submissions allowed\n"
            "• You can review existing submissions\n\n"
            "💡 Great for controlling when students can take the quiz!"
        )
    },
    
    # ========== QUIZ CODE & SHARING ==========
    {
        "tags": ["quiz code", "share code", "where is code", "find code", "code", "share quiz"],
        "q": "Where do I find the quiz code?",
        "a": (
            "🔑 **Finding the Quiz Code:**\n\n"
            "The code is shown on:\n"
            "• Quiz card: **Code: ABC123**\n"
            "• Quiz detail page header\n\n"
            "**Sharing with Students:**\n"
            "• Tell them the code verbally\n"
            "• Write it on the board\n"
            "• Share via class group\n"
            "• Use the QR code feature!\n\n"
            "Students enter this code on their dashboard to access the quiz."
        )
    },
    {
        "tags": ["qr code", "scan", "qr", "barcode", "scan code"],
        "q": "How do I use the QR code feature?",
        "a": (
            "📱 **Using QR Codes:**\n\n"
            "1. Open the quiz\n"
            "2. Click **Show QR Code**\n"
            "3. Display it on screen/projector\n"
            "4. Students scan with their phone camera\n\n"
            "Benefits:\n"
            "• No typing quiz code\n"
            "• Fast classroom access\n"
            "• Works with any QR scanner app\n\n"
            "💡 Students will be prompted to log in if not already!"
        )
    },
    
    # ========== STUDENTS & SUBMISSIONS ==========
    {
        "tags": ["student access", "login", "account required", "solve quiz", "student account", "student signup"],
        "q": "Can students access quizzes without an account?",
        "a": (
            "👤 **Student Account Requirements:**\n\n"
            "No, students MUST create an account and log in.\n\n"
            "**Student Sign Up Process:**\n"
            "1. Go to the app homepage\n"
            "2. Click **Student Sign Up**\n"
            "3. Fill in their details:\n"
            "   • Name, University ID\n"
            "   • Major, Section, City\n"
            "4. Create username & password\n\n"
            "This allows tracking individual submissions and grades."
        )
    },
    {
        "tags": ["view submissions", "see answers", "student responses", "results", "grades", "submissions"],
        "q": "How do I view student submissions?",
        "a": (
            "📊 **Viewing Submissions:**\n\n"
            "1. Open the quiz\n"
            "2. Click **View Submissions**\n"
            "3. See list of all students who submitted\n\n"
            "For each submission you can see:\n"
            "• Student name & ID\n"
            "• Score (auto-graded questions)\n"
            "• Submission time\n"
            "• Click to view detailed answers\n\n"
            "💡 Click **View & Grade** to grade file uploads!"
        )
    },
    {
        "tags": ["grade", "grading", "file upload grade", "mark", "score file", "grade file", "feedback"],
        "q": "How do I grade file upload submissions?",
        "a": (
            "✏️ **Grading File Uploads:**\n\n"
            "1. Open quiz → **View Submissions**\n"
            "2. Click **View & Grade** on a submission\n"
            "3. Click **👁️ Preview** to view the student's file\n"
            "4. Enter a **Grade** (e.g., A, B+, 85%)\n"
            "5. Add **Feedback Comments**\n"
            "6. Optionally upload a **Corrected File**\n"
            "7. Click **💾 Save Grading**\n\n"
            "Students will see your grade and feedback on their dashboard!"
        )
    },
    {
        "tags": ["extra attempt", "retry", "retake", "another try", "allow retry", "more attempts"],
        "q": "How do I allow a student to retake a quiz?",
        "a": (
            "🔄 **Allowing Extra Attempts:**\n\n"
            "1. Open quiz → **View Submissions**\n"
            "2. Find the student\n"
            "3. Click **Allow Extra Attempt**\n\n"
            "This gives them one more try. You can click multiple times for more attempts.\n\n"
            "💡 Useful for students who had technical issues!"
        )
    },
    
    # ========== EXPORT & REPORTS ==========
    {
        "tags": ["export", "excel", "download", "submissions file", "xlsx", "spreadsheet", "download grades"],
        "q": "How do I export submissions to Excel?",
        "a": (
            "📥 **Exporting to Excel:**\n\n"
            "**Option 1: Single Quiz**\n"
            "• Open quiz → Click **Export Excel**\n\n"
            "**Option 2: Entire Folder**\n"
            "• On folder card → Click **📥 Export**\n\n"
            "The Excel file includes:\n"
            "• Student full name\n"
            "• University ID\n"
            "• Section\n"
            "• Score/Grade\n"
            "• Submission time\n\n"
            "💡 Great for uploading to your university's grading system!"
        )
    },
    {
        "tags": ["analytics", "statistics", "performance", "weak topics", "ai analysis", "reports"],
        "q": "How do I see analytics and weak topics?",
        "a": (
            "📈 **Learning Analytics:**\n\n"
            "1. Open a **Subject Folder**\n"
            "2. Click **📊 Analytics**\n\n"
            "You'll see:\n"
            "• Most difficult questions (high error rates)\n"
            "• Student performance overview\n"
            "• AI-powered topic analysis\n\n"
            "Click **Analyze with AI** to get:\n"
            "• Weak topic identification\n"
            "• Root cause analysis\n"
            "• Teaching recommendations\n\n"
            "💡 Requires OpenAI API key in settings!"
        )
    },
    {
        "tags": ["flashcards", "ai flashcards", "student flashcards", "saved flashcards", "revision cards"],
        "q": "How do AI flashcards work for students?",
        "a": (
            "📇 **AI Flashcards for Students:**\n\n"
            "After a student submits a quiz, Quizfy can generate a saved revision set tied to that submission.\n\n"
            "What students get:\n"
            "• **Exactly 10 flashcards**\n"
            "• Each card is based on weak topics or mistakes\n"
            "• Cards are saved with the solved quiz\n\n"
            "Where students see them:\n"
            "• **Quiz Result** page\n"
            "• **Submission Details** page\n"
            "• **Saved Feedback** screen from the student dashboard\n\n"
            "💡 These flashcards are meant for quick review before trying the practice questions."
        )
    },
    {
        "tags": ["practice questions", "mcq practice", "true false practice", "ai questions", "saved practice", "7 questions"],
        "q": "What practice questions do students get after a quiz?",
        "a": (
            "✏️ **Saved AI Practice Questions:**\n\n"
            "After submission, students receive a saved practice set linked to that quiz attempt.\n\n"
            "The current practice set includes:\n"
            "• **Exactly 7 questions**\n"
            "• **MCQ and True/False only**\n"
            "• A fixed mix of **4 MCQ** and **3 True/False**\n\n"
            "Students can answer them on the feedback page, submit, and immediately see:\n"
            "• The correct answer\n"
            "• Whether their answer was correct\n"
            "• The explanation for the answer\n\n"
            "💡 Their practice answers are saved with the same quiz submission."
        )
    },
    {
        "tags": ["saved feedback", "student feedback", "ai feedback", "open feedback", "saved ai study set"],
        "q": "Where can students find their saved AI feedback?",
        "a": (
            "💡 **Where Students Find Saved AI Feedback:**\n\n"
            "Students can open their saved AI study materials from multiple places:\n\n"
            "1. **Quiz Result** right after submitting\n"
            "2. **Student Dashboard** from the quiz history card\n"
            "3. **Submission Details** for that solved quiz\n\n"
            "The saved feedback page includes:\n"
            "• Weak topics\n"
            "• 10 flashcards\n"
            "• 7 AI practice questions\n"
            "• Correct answers and explanations after submitting practice\n\n"
            "💡 Everything is stored per submission, so each solved quiz keeps its own study set."
        )
    },
    {
        "tags": ["students do after submit", "after quiz submission", "after students submit", "review answers", "answer explanation"],
        "q": "What happens after students submit a quiz?",
        "a": (
            "✅ **After Students Submit:**\n\n"
            "Students are taken to their result and review flow.\n\n"
            "They can:\n"
            "• Review their quiz answers\n"
            "• See the correct answers for the quiz\n"
            "• Open their saved AI feedback\n"
            "• Study 10 flashcards\n"
            "• Solve 7 saved practice questions\n"
            "• See the correct answer and explanation for each practice question after submitting\n\n"
            "💡 This makes the quiz attempt useful as both assessment and revision."
        )
    },
    
    # ========== ACCOUNT & SECURITY ==========
    {
        "tags": ["change password", "password", "security", "update password", "new password"],
        "q": "How do I change my password?",
        "a": (
            "🔐 **Changing Your Password:**\n\n"
            "1. Click **Change Password** (top of dashboard)\n"
            "2. Enter your current password\n"
            "3. Enter new password (twice)\n"
            "4. Click **Change Password**\n\n"
            "💡 Use a strong password with letters, numbers, and symbols!"
        )
    },
    {
        "tags": ["logout", "sign out", "log out"],
        "q": "How do I log out?",
        "a": (
            "👋 **Logging Out:**\n\n"
            "Click the **Logout** link in the navigation.\n\n"
            "💡 Always log out when using shared computers!"
        )
    },
    
    # ========== TROUBLESHOOTING ==========
    {
        "tags": ["student can't access", "student problem", "quiz not working", "error", "issue", "problem"],
        "q": "A student can't access the quiz. What should I check?",
        "a": (
            "🔧 **Troubleshooting Student Access:**\n\n"
            "Check these things:\n\n"
            "1. **Is the quiz active?**\n"
            "   • Make sure you clicked Start Quiz\n\n"
            "2. **Has the due date passed?**\n"
            "   • Check quiz settings\n\n"
            "3. **Did they use the correct code?**\n"
            "   • Codes are case-insensitive\n\n"
            "4. **Are they logged in?**\n"
            "   • Students need an account\n\n"
            "5. **Have they already submitted?**\n"
            "   • Default is 1 attempt - allow extra if needed"
        )
    },
    {
        "tags": ["missing submission", "didn't submit", "lost answers", "no submission"],
        "q": "A student says they submitted but I don't see it?",
        "a": (
            "🔍 **Finding Missing Submissions:**\n\n"
            "1. Check the submissions list carefully (sorted by time)\n"
            "2. The student may have:\n"
            "   • Started but not clicked Submit\n"
            "   • Had internet issues\n"
            "   • Used a different account\n\n"
            "**Solutions:**\n"
            "• Allow an extra attempt for them\n"
            "• Have them try again\n"
            "• Check if timer auto-submitted (shows in submission time)"
        )
    },
    
    # ========== TIPS & BEST PRACTICES ==========
    {
        "tags": ["tips", "best practices", "advice", "help", "how to use", "guide"],
        "q": "Any tips for using Quizfy effectively?",
        "a": (
            "💡 **Pro Tips:**\n\n"
            "**Organization:**\n"
            "• Create folders by subject/course\n"
            "• Use clear quiz titles with dates\n\n"
            "**Quizzes:**\n"
            "• Test your quiz before sharing\n"
            "• Use QR codes for fast access\n"
            "• Set appropriate time limits\n\n"
            "**Grading:**\n"
            "• Export to Excel regularly\n"
            "• Use analytics to find weak areas\n"
            "• Provide detailed feedback on file uploads\n\n"
            "**AI Study Support:**\n"
            "• Encourage students to open **Saved Feedback** after every quiz\n"
            "• Use the flashcards for quick revision\n"
            "• Ask students to complete the 7 saved practice questions for reinforcement\n\n"
            "**Students:**\n"
            "• Share the code clearly\n"
            "• Remind them to click Submit!"
        )
    },
    {
        "tags": ["hello", "hi", "hey", "greetings"],
        "q": "Hello",
        "a": (
            "👋 **Hello!** I'm your Quizfy Assistant.\n\n"
            "I can help you with:\n"
            "• Creating and managing quizzes\n"
            "• Organizing subject folders\n"
            "• Grading student submissions\n"
            "• Exporting grades to Excel\n"
            "• Using analytics features\n"
            "• Understanding saved AI flashcards and practice feedback\n\n"
            "Just ask me a question!"
        )
    },
    {
        "tags": ["thank", "thanks", "thank you"],
        "q": "Thank you",
        "a": (
            "😊 You're welcome! Happy to help.\n\n"
            "If you have more questions, just ask!"
        )
    },
]

def _normalize(text: str) -> str:
    text = (text or "").lower().strip()
    text = re.sub(r"\s+", " ", text)
    return text

def _score_match(query: str, item: dict) -> float:
    """Score how well a query matches a KB item"""
    query_norm = _normalize(query)
    query_words = set(query_norm.split())
    
    # Remove common filler words for matching
    filler_words = {"the", "a", "an", "is", "are", "can", "i", "you", "how", "do", "to", "my", "this", "that", "what", "where", "when", "please", "help", "want", "need"}
    query_words = query_words - filler_words
    
    score = 0.0
    
    # Check tags (highest priority)
    for tag in item["tags"]:
        tag_norm = _normalize(tag)
        tag_words = set(tag_norm.split()) - filler_words
        
        # Exact tag match in query
        if tag_norm in query_norm:
            score += 15.0
        
        # Query contains tag
        if query_norm in tag_norm:
            score += 10.0
        
        # Word overlap
        overlap = len(query_words & tag_words)
        if overlap > 0:
            score += overlap * 3.0
    
    # Check question text
    q_norm = _normalize(item["q"])
    q_words = set(q_norm.split()) - filler_words
    overlap = len(query_words & q_words)
    score += overlap * 2.0
    
    # Fuzzy matching for typos
    for qw in query_words:
        if len(qw) < 3:
            continue
        for tag in item["tags"]:
            tag_words = _normalize(tag).split()
            matches = difflib.get_close_matches(qw, tag_words, n=1, cutoff=0.75)
            if matches:
                score += 2.0
    
    return score

def _best_answer(message: str) -> str:
    msg = message.strip()
    
    if not msg:
        return "👋 Hi! Ask me anything about using Quizfy. For example:\n• How do I create a quiz?\n• How do I grade submissions?\n• How do I export to Excel?"
    
    # Score all KB items
    scores = []
    for item in TEACHER_KB:
        score = _score_match(msg, item)
        if score > 0:
            scores.append((score, item))
    
    # Sort by score (highest first)
    scores.sort(key=lambda x: x[0], reverse=True)
    
    if scores and scores[0][0] >= 3.0:
        return scores[0][1]["a"]
    
    # Fallback with suggestions
    return (
        "🤔 I'm not sure about that. Here are some things I can help with:\n\n"
        "**Quizzes:**\n"
        "• How do I create a quiz?\n"
        "• How do I add questions?\n"
        "• How do I set a time limit?\n\n"
        "**Grading:**\n"
        "• How do I view submissions?\n"
        "• How do I grade file uploads?\n"
        "• How do I export to Excel?\n\n"
        "**AI Feedback:**\n"
        "• How do AI flashcards work for students?\n"
        "• What practice questions do students get after a quiz?\n"
        "• Where can students find their saved AI feedback?\n\n"
        "**Organization:**\n"
        "• How do I create folders?\n"
        "• How do I use analytics?\n\n"
        "Try asking one of these questions!"
    )

@staff_required
@require_POST
def teacher_help_bot(request):
    try:
        data = json.loads(request.body.decode("utf-8"))
    except Exception:
        data = {}

    message = (data.get("message") or "").strip()
    if not message:
        return JsonResponse({"reply": "Type your question and I’ll help."})

    reply = _best_answer(message)
    return JsonResponse({"reply": reply})


@student_required
def enter_quiz(request):
    form = EnterQuizForm(request.POST or None)

    if request.method == "POST" and form.is_valid():
        quiz_code = form.cleaned_data["quiz_code"].strip().upper()

        quiz = Quiz.objects.filter(code=quiz_code).first()
        if not quiz:
            messages.error(request, "Invalid quiz code.")
            return render(request, "quizzes/enter_quiz.html", {"form": form})

        return redirect("take_quiz", quiz_code=quiz_code)

    return render(request, "quizzes/enter_quiz.html", {"form": form})

@staff_required
@require_POST
def allow_extra_attempt(request, quiz_id, student_id):
    quiz = get_object_or_404(Quiz, id=quiz_id, teacher=request.user)
    student = get_object_or_404(User, id=student_id)

    perm, _ = QuizAttemptPermission.objects.get_or_create(
        quiz=quiz, student_user=student,
        defaults={"allowed_attempts": 1}
    )
    perm.allowed_attempts += 1
    perm.save()

    messages.success(request, f"Extra attempt allowed for {student.username}.")
    return redirect("quiz_submissions", quiz_id=quiz.id)

@staff_required
def edit_quiz_settings(request, quiz_id):
    quiz = get_object_or_404(Quiz, id=quiz_id, teacher=request.user)

    form = QuizSettingsForm(request.POST or None, instance=quiz)

    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "Quiz settings updated.")
        return redirect("teacher_quiz_detail", quiz_id=quiz.id)

    return render(request, "quizzes/edit_quiz_settings.html", {
        "quiz": quiz,
        "form": form
    })


def quiz_join(request, quiz_code):
    """
    Join Quiz page - shown when QR code is scanned.
    Pre-fills the quiz code and shows quiz info.
    Redirects to login/signup if needed, then to quiz.
    """
    quiz = get_object_or_404(Quiz, code=quiz_code.upper())
    
    # Check if quiz is available
    if not quiz.can_start():
        messages.error(request, "This quiz is currently closed.")
        return render(request, "quizzes/quiz_join.html", {
            "quiz": quiz,
            "quiz_code": quiz.code,
            "is_closed": True,
        })
    
    if request.user.is_authenticated and request.user.is_staff:
        messages.info(request, "You're logged in as a teacher. Students can join this quiz.")
    
    return render(request, "quizzes/quiz_join.html", {
        "quiz": quiz,
        "quiz_code": quiz.code,
        "is_closed": False,
    })


def quiz_scan_redirect(request, quiz_code):
    """
    Endpoint specifically for QR code scanning.
    Checks authentication and redirects to login if needed,
    then to the quiz with proper next parameter.
    """
    join_url = reverse("quiz_join", kwargs={"quiz_code": quiz_code})

    if not request.user.is_authenticated:
        return redirect(join_url)

    if not hasattr(request.user, "student_profile"):
        messages.error(request, "Only students can take quizzes.")
        return redirect("student_dashboard")

    return redirect(join_url)

@ensure_csrf_cookie
def take_quiz(request, quiz_code):
    # Check if user is authenticated and is a student
    if not request.user.is_authenticated:
        # Redirect to login with next parameter pointing back to this quiz
        from django.urls import reverse
        next_url = reverse('take_quiz', kwargs={'quiz_code': quiz_code})
        return redirect(f"/student/login/?next={next_url}")
    
    if not hasattr(request.user, "student_profile"):
        messages.error(request, "Only students can take quizzes.")
        return redirect("student_dashboard")
    
    quiz = get_object_or_404(Quiz, code=quiz_code.upper())

    # ✅ 1) Teacher stopped quiz OR due date passed
    if not quiz.can_start():
        messages.error(request, "This quiz is closed.")
        return redirect("student_dashboard")

    # ✅ 2) Attempt limit (default = 1)
    perm = QuizAttemptPermission.objects.filter(
        quiz=quiz, student_user=request.user
    ).first()
    allowed = perm.allowed_attempts if perm else 1

    attempts_used = Submission.objects.filter(
        quiz=quiz,
        student_user=request.user,
        is_submitted=True
    ).count()

    if attempts_used >= allowed:
        messages.error(request, "You already submitted this quiz.")
        return redirect("student_dashboard")

    questions = quiz.questions.all().order_by("id")

    # ✅ 3) Get or create an in-progress submission (so refresh won’t create new one)
    submission = Submission.objects.filter(
        quiz=quiz,
        student_user=request.user,
        is_submitted=False
    ).order_by("-id").first()

    if not submission:
        sp = request.user.student_profile
        # For file upload quizzes, total is 0 (manual grading)
        total_questions = 0 if quiz.quiz_type == 'file_upload' else questions.count()
        submission = Submission.objects.create(
            quiz=quiz,
            student_user=request.user,
            student_name=f"{sp.first_name} {sp.second_name} {sp.third_name}",
            score=0,
            total=total_questions,
            started_at=timezone.now(),
            is_submitted=False,
            attempt_no=attempts_used + 1,
        )
        logger.info(
            "quiz_submission_created: "
            "submission_id=%d student_user=%s student_name=%s quiz=%s total_questions=%d",
            submission.id,
            request.user.id,
            submission.student_name,
            quiz.code,
            total_questions,
        )
    else:
        # ✅ IMPORTANT PATCH: old rows may have started_at = NULL
        if submission.started_at is None:
            submission.started_at = timezone.now()
            submission.save(update_fields=["started_at"])

    # ✅ 4) Timer in minutes -> seconds
    remaining_seconds = None
    if quiz.duration_minutes:
        duration_seconds = quiz.duration_minutes * 60
        elapsed = (timezone.now() - submission.started_at).total_seconds()
        remaining_seconds = max(0, int(duration_seconds - elapsed))

        # ✅ If time is finished (GET or POST), finalize NOW
        if remaining_seconds <= 0:
            return _finalize_submission(request, quiz, submission, questions)

    # ✅ 5) POST = finalize (manual submit OR timer submits the form)
    if request.method == "POST":
        return _finalize_submission(request, quiz, submission, questions)

    # ✅ 6) GET = show quiz
    student_profile = request.user.student_profile
    return render(request, "quizzes/take_quiz.html", {
        "quiz": quiz,
        "questions": questions,
        "remaining_seconds": remaining_seconds,
        "submission": submission,
        "secure_quiz_enabled": True,
        "secure_quiz_violation_limit": SECURE_QUIZ_AUTO_SUBMIT_THRESHOLD,
        "student_display_name": submission.student_name,
        "student_university_id": student_profile.university_id,
    })

@transaction.atomic
def _finalize_submission(request, quiz, submission, questions=None):
    from django.contrib import messages  # Import at top of function
    
    # If already submitted, go to result
    if submission.is_submitted:
        return redirect("quiz_result", quiz_code=quiz.code, submission_id=submission.id)

    # Teacher stopped it / due date passed mid-attempt
    if not quiz.can_start():
        messages.error(request, "Quiz was closed by the teacher.")
        return redirect("student_dashboard")

    # Handle file upload quiz type
    if quiz.quiz_type == 'file_upload':
        uploaded_file = request.FILES.get('file')
        if uploaded_file:
            # Validate file
            max_size = 10 * 1024 * 1024  # 10MB
            allowed_extensions = ['pdf', 'jpg', 'jpeg', 'png']
            ext = uploaded_file.name.split('.')[-1].lower()
            
            if uploaded_file.size > max_size:
                messages.error(request, "File size must be under 10MB")
                return redirect("take_quiz", quiz_code=quiz.code)
            
            if ext not in allowed_extensions:
                messages.error(request, "Only PDF, JPG, and PNG files are allowed")
                return redirect("take_quiz", quiz_code=quiz.code)
            
            # Save file submission
            FileSubmission.objects.create(
                submission=submission,
                file=uploaded_file,
                file_name=uploaded_file.name
            )
            
            submission.score = 0  # Will be graded manually
            submission.total = 0
            submission.is_submitted = True
            submission.submitted_at = timezone.now()
            submission.save()
            
            messages.success(request, "Your file has been submitted successfully!")
            return redirect("quiz_result", quiz_code=quiz.code, submission_id=submission.id)
        else:
            messages.error(request, "Please upload a file")
            return redirect("take_quiz", quiz_code=quiz.code)

    # If questions not provided, load them
    if questions is None:
        questions = quiz.questions.all().order_by("id")

    # Remove old answers (if re-submit happens)
    submission.answers.all().delete()

    score = 0
    total = 0  # Count non-file-upload questions

    for q in questions:
        if q.question_type == 'file_upload':
            # Handle file upload question
            uploaded_file = request.FILES.get(f"file_{q.id}")
            if uploaded_file:
                # Validate file
                max_size = 10 * 1024 * 1024  # 10MB
                allowed_extensions = ['pdf', 'jpg', 'jpeg', 'png']
                ext = uploaded_file.name.split('.')[-1].lower()
                
                if uploaded_file.size > max_size:
                    messages.error(request, f"File for Q{q.id} must be under 10MB")
                    return redirect("take_quiz", quiz_code=quiz.code)
                
                if ext not in allowed_extensions:
                    messages.error(request, f"Only PDF, JPG, and PNG files are allowed")
                    return redirect("take_quiz", quiz_code=quiz.code)
                
                # Save file submission linked to the question
                FileSubmission.objects.create(
                    submission=submission,
                    question=q,
                    file=uploaded_file,
                    file_name=uploaded_file.name
                )
            
            # File upload questions are graded manually (don't count toward auto-score)
            Answer.objects.create(
                submission=submission,
                question=q,
                selected=None,
                is_correct=False,  # Will be graded manually
            )
        else:
            # Handle multiple choice / true-false
            total += 1
            selected = request.POST.get(f"question_{q.id}")
            selected_int = int(selected) if selected else None

            is_correct = (selected_int == q.correct_option)
            if is_correct:
                score += 1

            Answer.objects.create(
                submission=submission,
                question=q,
                selected=selected_int,
                is_correct=is_correct,
            )

    submission.score = score
    submission.total = total
    submission.is_submitted = True
    submission.submitted_at = timezone.now()
    
    logger.info(
        "quiz_submission_finalizing: "
        "submission_id=%d student_user=%s student_name=%s quiz=%s score=%d/%d is_submitted=%s",
        submission.id,
        submission.student_user_id,
        submission.student_name,
        submission.quiz.code,
        score,
        total,
        submission.is_submitted,
    )
    
    submission.save()
    
    logger.info(
        "quiz_submission_saved: "
        "submission_id=%d student_user=%s is_submitted=%s",
        submission.id,
        submission.student_user_id,
        submission.is_submitted,
    )

    # Generate personalized feedback (runs synchronously for now, can be async later)
    try:
        _generate_student_feedback(submission)
    except Exception as e:
        logger.error(
            "Failed to generate feedback for submission_id=%d: %s",
            submission.id,
            str(e),
            exc_info=True,
        )
        # Don't block quiz submission if feedback generation fails

    return redirect("quiz_result", quiz_code=quiz.code, submission_id=submission.id)


def _generate_student_feedback(submission):
    """
    Generate AI-powered personalized feedback for a student's quiz submission.
    
    This function:
    1. Creates StudentFeedback record
    2. Uses LLM to analyze mistakes and generate feedback
    3. Stores feedback JSON (weak topics, flashcards, practice questions)
    4. Handles errors gracefully without blocking quiz submission
    """
    from quizzes.models import StudentFeedback
    from quizzes.services.student_feedback_generator import StudentFeedbackGenerator
    
    # Create or get feedback record
    feedback, created = StudentFeedback.objects.get_or_create(
        submission=submission,
        defaults={
            'weak_topics': [],
            'flashcards': [],
            'practice_questions': [],
        }
    )
    
    if not created and StudentFeedbackGenerator.has_expected_feedback_shape(feedback):
        # Feedback already generated, skip
        logger.info(
            "Feedback already generated for submission_id=%d, skipping",
            submission.id,
        )
        return
    
    try:
        # Generate feedback using LLM
        generator = StudentFeedbackGenerator()
        feedback_data = generator.generate_feedback(submission)
        
        # Update feedback record with results
        feedback.weak_topics = feedback_data.get('weak_topics', [])
        feedback.flashcards = feedback_data.get('flashcards', [])
        feedback.practice_questions = feedback_data.get('practice_questions', [])
        feedback.generation_error = None
        feedback.save()
        
        logger.info(
            "Feedback generated successfully for submission_id=%d student=%s",
            submission.id,
            submission.student_name,
        )
        
    except Exception as e:
        # Store error but don't crash
        error_msg = f"{type(e).__name__}: {str(e)}"
        feedback.generation_error = error_msg
        feedback.save()
        
        logger.error(
            "Failed to generate feedback for submission_id=%d: %s",
            submission.id,
            error_msg,
            exc_info=True,
        )


@student_required
@require_POST
def log_quiz_security_violation(request, quiz_code, submission_id):
    quiz = get_object_or_404(Quiz, code=quiz_code.upper())
    submission = get_object_or_404(
        Submission,
        id=submission_id,
        quiz=quiz,
        student_user=request.user,
    )

    try:
        payload = json.loads(request.body.decode("utf-8"))
    except Exception:
        payload = {}

    violation_type = str(payload.get("violation_type") or "").strip().upper()
    details = str(payload.get("details") or "").strip()
    allowed_types = {
        choice[0] for choice in QuizSecurityViolation.VIOLATION_TYPE_CHOICES
    }

    if violation_type not in allowed_types:
        return JsonResponse({"ok": False, "error": "Invalid violation type."}, status=400)

    violation = QuizSecurityViolation.objects.create(
        student=request.user,
        quiz=quiz,
        attempt=submission,
        violation_type=violation_type,
        details=details or None,
    )
    violation_count = submission.security_violations.count()
    should_auto_submit = violation_count >= SECURE_QUIZ_AUTO_SUBMIT_THRESHOLD and not submission.is_submitted
    result_url = None

    if should_auto_submit:
        reason_text = (
            f"Your quiz was automatically submitted after {violation_count} recorded security violations. "
            f"Last recorded action: {violation.get_violation_type_display()}."
        )
        submission = _finalize_security_forced_submission(quiz, submission, reason=reason_text)
        result_url = reverse(
            "quiz_result",
            kwargs={"quiz_code": quiz.code, "submission_id": submission.id},
        )

    return JsonResponse(
        {
            "ok": True,
            "violation_id": violation.id,
            "violation_count": violation_count,
            "security_status": _security_status_from_count(violation_count),
            "should_auto_submit": should_auto_submit,
            "result_url": result_url,
        }
    )


def _feedback_has_saved_study_set(feedback):
    if not feedback:
        return False
    return bool(
        (feedback.weak_topics or [])
        or (feedback.flashcards or [])
        or (feedback.practice_questions or [])
    )


def _get_submission_feedback(submission):
    try:
        return submission.feedback
    except Exception:
        return None


def _normalize_practice_answer(value):
    return str(value or "").strip()


def _security_status_from_count(violation_count):
    count = int(violation_count or 0)
    if count >= SECURE_QUIZ_HIGH_RISK_THRESHOLD:
        return "High Risk"
    if count >= SECURE_QUIZ_SUSPICIOUS_THRESHOLD:
        return "Suspicious"
    return "Clean"


def _attach_security_summary(submission, prefetched_violations=None):
    violations = list(prefetched_violations if prefetched_violations is not None else submission.security_violations.all())
    violation_count = len(violations)
    submission.security_violation_count = violation_count
    submission.security_status = _security_status_from_count(violation_count)
    submission.security_violations_list = violations
    return submission


def _finalize_security_forced_submission(quiz, submission, reason=None):
    """Force-submit a quiz attempt after repeated security violations."""
    if submission.is_submitted:
        return submission

    questions = quiz.questions.all().order_by("id")
    submission.answers.all().delete()
    submission.file_submissions.all().delete()

    score = 0
    total = 0

    for question in questions:
        if question.question_type == "file_upload":
            Answer.objects.create(
                submission=submission,
                question=question,
                selected=None,
                is_correct=False,
            )
            continue

        total += 1
        Answer.objects.create(
            submission=submission,
            question=question,
            selected=None,
            is_correct=False,
        )

    submission.score = score
    submission.total = total
    submission.is_submitted = True
    submission.submitted_at = timezone.now()
    submission.security_forced_submit_reason = reason or (
        f"Auto-submitted after reaching {SECURE_QUIZ_AUTO_SUBMIT_THRESHOLD} recorded security violations."
    )
    submission.save(update_fields=[
        "score",
        "total",
        "is_submitted",
        "submitted_at",
        "security_forced_submit_reason",
    ])
    _generate_student_feedback(submission)
    return submission


def _grade_saved_practice_questions(practice_questions, submitted_answers):
    attempts = []
    for index, question in enumerate(practice_questions):
        selected_answer = _normalize_practice_answer(
            submitted_answers.get(f"practice_answer_{index}")
        )
        correct_answer = _normalize_practice_answer(question.get("correct_answer"))
        attempts.append(
            {
                "question_index": index,
                "selected_answer": selected_answer,
                "correct_answer": correct_answer,
                "is_correct": selected_answer.lower() == correct_answer.lower(),
            }
        )
    return attempts


def _build_practice_question_cards(practice_questions, saved_attempts=None, draft_answers=None):
    attempt_map = {}
    for attempt in saved_attempts or []:
        if not isinstance(attempt, dict):
            continue
        try:
            question_index = int(attempt.get("question_index"))
        except (TypeError, ValueError):
            continue
        attempt_map[question_index] = attempt

    cards = []
    for index, question in enumerate(practice_questions or []):
        selected_answer = _normalize_practice_answer(
            (draft_answers or {}).get(f"practice_answer_{index}")
        )
        saved_attempt = attempt_map.get(index)
        if saved_attempt:
            selected_answer = _normalize_practice_answer(saved_attempt.get("selected_answer"))

        option_rows = []
        for option_index, option in enumerate(question.get("options") or []):
            option_value = str(option)
            option_rows.append(
                {
                    "label": chr(65 + option_index) if question.get("type") == "mcq" else "",
                    "value": option_value,
                    "is_selected": selected_answer == option_value,
                    "is_correct": _normalize_practice_answer(question.get("correct_answer")).lower() == option_value.strip().lower(),
                }
            )

        card = dict(question)
        card.update(
            {
                "index": index,
                "field_name": f"practice_answer_{index}",
                "selected_answer": selected_answer,
                "option_rows": option_rows,
                "attempt": saved_attempt,
            }
        )
        cards.append(card)

    return cards


@staff_required
@require_POST
def toggle_quiz_active(request, quiz_id):
    quiz = get_object_or_404(Quiz, id=quiz_id, teacher=request.user)

    if request.method == "POST":
        quiz.is_active = not quiz.is_active
        quiz.save()

    return redirect("teacher_quiz_detail", quiz_id=quiz.id)

@staff_required
@require_POST
def allow_extra_attempt(request, quiz_id, student_id):
    quiz = get_object_or_404(Quiz, id=quiz_id, teacher=request.user)
    student = get_object_or_404(User, id=student_id)

    perm, created = QuizAttemptPermission.objects.get_or_create(
        quiz=quiz,
        student_user=student,
        defaults={"allowed_attempts": 1}
    )

    if not created:
        perm.allowed_attempts += 1
        perm.save()

    messages.success(request, f"Extra attempt allowed for {student.username}.")
    return redirect("quiz_submissions", quiz_id=quiz.id)


def quiz_result(request, quiz_code, submission_id):
    quiz = get_object_or_404(Quiz, code=quiz_code)

    # ✅ scope submission to the quiz to avoid mismatches
    submission = get_object_or_404(Submission, id=submission_id, quiz=quiz)

    # ✅ safe student display
    student_name = submission.student_name or ""
    university_id = ""

    if submission.student_user_id:
        try:
            u = submission.student_user
            sp = getattr(u, "student_profile", None)
            if sp:
                student_name = f"{sp.first_name} {sp.second_name} {sp.third_name}".strip() or student_name
                university_id = sp.university_id or ""
            else:
                # fallback to username if no profile
                student_name = u.username or student_name
        except User.DoesNotExist:
            pass

    # Get answers for the submission
    answers = submission.answers.select_related('question').order_by('id')
    
    # Get file submissions
    file_submissions = submission.file_submissions.select_related('question').all()
    
    # Check if feedback is available
    feedback = None
    feedback_ready = False
    feedback = _get_submission_feedback(submission)
    feedback_ready = _feedback_has_saved_study_set(feedback)

    context = {
        "quiz": quiz,
        "submission": submission,
        "student_name": student_name,
        "university_id": university_id,
        "answers": answers,
        "file_submissions": file_submissions,
        "feedback": feedback,
        "feedback_ready": feedback_ready,
        "security_forced_submit": bool(submission.security_forced_submit_reason),
    }
    return render(request, "quizzes/quiz_result.html", context)


def student_feedback(request, quiz_code, submission_id):
    """
    Display personalized learning feedback for a student's quiz submission.
    
    Feedback is generated automatically after quiz submission and includes:
    - Weak topics identified from incorrect answers
    - Flashcards for revision
    - Adaptive practice questions with solutions
    """
    from quizzes.models import StudentFeedback
    from quizzes.services.student_feedback_generator import StudentFeedbackGenerator

    quiz = get_object_or_404(Quiz, code=quiz_code)
    submission = get_object_or_404(Submission, id=submission_id, quiz=quiz)
    
    # Ensure student can only see their own feedback
    if submission.student_user and submission.student_user != request.user:
        messages.error(request, "You do not have access to this feedback.")
        return redirect("home")
    
    draft_answers = {}

    # Mark feedback as viewed
    try:
        feedback = submission.feedback
        if _feedback_has_saved_study_set(feedback) and not StudentFeedbackGenerator.has_expected_feedback_shape(feedback):
            _generate_student_feedback(submission)
            feedback.refresh_from_db()
        feedback.mark_as_viewed()
    except StudentFeedback.DoesNotExist:
        messages.warning(request, "Feedback is still being generated. Please try again in a moment.")
        return redirect("quiz_result", quiz_code=quiz_code, submission_id=submission_id)

    if not _feedback_has_saved_study_set(feedback) and not feedback.generation_error:
        messages.warning(request, "Feedback is still being generated. Please try again in a moment.")
        return redirect("quiz_result", quiz_code=quiz_code, submission_id=submission_id)
    
    # Handle any generation errors
    if feedback.generation_error:
        messages.error(
            request,
            f"Feedback generation encountered an error: {feedback.generation_error}"
        )

    practice_questions = feedback.practice_questions or []
    if request.method == "POST" and practice_questions:
        draft_answers = {
            f"practice_answer_{index}": request.POST.get(f"practice_answer_{index}", "")
            for index in range(len(practice_questions))
        }
        missing_answers = [
            str(index + 1)
            for index, question in enumerate(practice_questions)
            if not _normalize_practice_answer(draft_answers.get(f"practice_answer_{index}"))
        ]
        if missing_answers:
            messages.error(
                request,
                "Please answer all 7 saved practice questions before submitting."
            )
        else:
            feedback.practice_question_attempts = _grade_saved_practice_questions(
                practice_questions,
                draft_answers,
            )
            feedback.practice_attempted_at = timezone.now()
            feedback.save(update_fields=["practice_question_attempts", "practice_attempted_at", "updated_at"])
            messages.success(request, "Your saved practice answers were checked and stored with this quiz.")
            return redirect("student_feedback", quiz_code=quiz_code, submission_id=submission_id)

    practice_question_cards = _build_practice_question_cards(
        practice_questions,
        saved_attempts=feedback.practice_question_attempts or [],
        draft_answers=draft_answers,
    )
    practice_attempts = feedback.practice_question_attempts or []
    practice_score = sum(1 for attempt in practice_attempts if attempt.get("is_correct"))

    # Parse JSON data for template
    context = {
        "quiz": quiz,
        "submission": submission,
        "student_name": submission.student_name,
        "feedback": feedback,
        "weak_topics": feedback.weak_topics or [],
        "flashcards": feedback.flashcards or [],
        "practice_questions": practice_question_cards,
        "practice_answers_saved": bool(practice_attempts),
        "practice_score": practice_score,
        "practice_total": len(practice_question_cards),
        "feedback_ready": _feedback_has_saved_study_set(feedback),
    }
    
    return render(request, "quizzes/student_feedback.html", context)


@transaction.atomic
def teacher_signup(request):
    if request.user.is_authenticated:
        return redirect("teacher_quizzes")
    form=TeacherSignupForm(request.POST or None)
    if request.method=='POST' and form.is_valid():
        user=form.save(commit=False)
        user.set_password(form.cleaned_data['password1'])
        user.is_staff=True
        user.save()
        
        login(request,user)
        messages.success(request,"Account created successfully.")
        return redirect("teacher_quizzes")
    return render(request,"quizzes/teacher_signup.html",{'form':form})

def teacher_login(request):
    if request.user.is_authenticated:
        # if already logged-in teacher, go dashboard; if student, kick them to student dashboard
        if request.user.is_staff:
            return redirect("teacher_quizzes")
        return redirect("student_dashboard")

    form = TeacherLoginForm(request.POST or None)

    if request.method == "POST" and form.is_valid():
        user = authenticate(
            request,
            username=form.cleaned_data["username"],
            password=form.cleaned_data["password"],
        )
        if user and user.is_staff:
            login(request, user)
            return redirect("teacher_quizzes")
        messages.error(request, "Invalid teacher credentials.")

    return render(request, "quizzes/teacher_login.html", {"form": form})

@staff_required
def teacher_quizzes(request):
    folders = SubjectFolder.objects.filter(teacher=request.user).order_by("name")
    ungrouped = (
        Quiz.objects.filter(teacher=request.user, folder__isnull=True)
        .annotate(
            assigned_count=Count("attempt_permissions", distinct=True),
            submitted_count=Count(
                "submissions",
                filter=Q(submissions__is_submitted=True),
                distinct=True,
            ),
        )
        .order_by("-created_at")
    )
    for quiz in ungrouped:
        quiz.bar_max = max(1, quiz.assigned_count, quiz.submitted_count)
    return render(request, "quizzes/teacher_folders.html", {
        "folders": folders,
        "ungrouped": ungrouped,
    })




@staff_required
def performance_analytics_dashboard(request):
    """
    Teacher-facing test page for the class-first PDF + Excel analytics workflow.
    Keeps all LLM and vector work on the backend.
    """
    context = {
        "analysis": None,
        "upload_summary": None,
    }

    def render_analytics_response():
        logger.info(
            "teacher_analytics_response_returned method=%s user_id=%s has_analysis=%s",
            request.method,
            request.user.id,
            bool(context.get("analysis")),
        )
        return render(request, "quizzes/performance_analytics.html", context)

    logger.info(
        "teacher_analytics_view_entered method=%s user_id=%s",
        request.method,
        request.user.id,
    )

    if request.method != "POST":
        return render_analytics_response()

    logger.info("teacher_analytics_csrf_passed user_id=%s", request.user.id)
    logger.info("teacher_analytics_post_received user_id=%s", request.user.id)

    pdf_file = request.FILES.get("pdf_file")
    excel_file = request.FILES.get("excel_file")
    logger.info(
        "teacher_analytics_files_present user_id=%s pdf=%s excel=%s",
        request.user.id,
        bool(pdf_file),
        bool(excel_file),
    )

    if not pdf_file or not excel_file:
        messages.error(request, "Upload both the assessment PDF and the class results file.")
        return render_analytics_response()

    def upload_extension(upload):
        if "." not in upload.name:
            return ""
        return upload.name.rsplit(".", 1)[-1].lower()

    if not pdf_file.name.lower().endswith(".pdf"):
        messages.error(request, "The assessment file must be a PDF.")
        return render_analytics_response()

    allowed_results = (".xlsx", ".xls", ".csv", ".tsv")
    if not excel_file.name.lower().endswith(allowed_results):
        messages.error(request, "The results file must be Excel, CSV, or TSV.")
        return render_analytics_response()

    def is_openai_rate_limit_error(exc):
        return exc.__class__.__name__ == "RateLimitError" or getattr(exc, "status_code", None) == 429

    def safe_analytics_error_message(exc):
        exc_name = exc.__class__.__name__
        status_code = getattr(exc, "status_code", None)
        if exc_name == "AuthenticationError" or status_code == 401:
            return "OpenAI authentication failed. Check the backend OPENAI_API_KEY in Render."
        if is_openai_rate_limit_error(exc):
            return "OpenAI rate limit or quota was reached. Please wait and try again, or check the backend OpenAI billing/rate limit settings."
        if exc_name == "BadRequestError" or status_code == 400:
            return "OpenAI could not process this PDF content. Try a smaller or clearer assessment PDF."
        if exc_name in {"APIConnectionError", "APITimeoutError", "TimeoutException"}:
            return "Could not reach OpenAI from the backend. Please try again."
        return "Analysis failed. Please check that the PDF has numbered questions and the Excel has student_id/student_name plus Q1, Q2, Q3 columns with 1 or 0 values."

    try:
        from performance_analytics.config import config
        from performance_analytics.models import AnalysisRequest
        from performance_analytics.services.analyzer import PerformanceAnalyzer
        from performance_analytics.services.excel_processor import ExcelProcessor
        from performance_analytics.services.pdf_processor import PDFProcessor
        from performance_analytics.services.topic_classifier import TopicClassifier
        from performance_analytics.services.vector_store import vector_store
        from performance_analytics.storage import store

        config.validate_startup()
        vector_store.configure()
        logger.info(
            "analytics_submit_started user_id=%s pdf_ext=%s results_ext=%s pdf_size=%d results_size=%d",
            request.user.id,
            upload_extension(pdf_file),
            upload_extension(excel_file),
            getattr(pdf_file, "size", 0) or 0,
            getattr(excel_file, "size", 0) or 0,
        )

        pdf_id = store.new_id("pdf")
        excel_id = store.new_id("excel")
        pdf_path = store.save_upload(pdf_id, pdf_file.name, pdf_file.read())
        excel_path = store.save_upload(excel_id, excel_file.name, excel_file.read())

        extracted = PDFProcessor().extract_questions(pdf_path, pdf_id)
        if not extracted:
            raise ValueError("No questions could be extracted from the PDF.")
        logger.info(
            "analytics_pdf_extracted user_id=%s question_count=%d",
            request.user.id,
            len(extracted),
        )

        classified = TopicClassifier().classify_batch(extracted, subject=None)
        vector_store.ingest_questions(classified)
        store.save_questions(pdf_id, classified)
        logger.info(
            "analytics_topics_classified user_id=%s question_count=%d",
            request.user.id,
            len(classified),
        )

        results = ExcelProcessor().normalize_results(excel_path)
        if not results:
            raise ValueError("No student result rows could be read from the results file.")
        store.save_results(excel_id, results)
        student_count = len({result.student_id for result in results})
        logger.info(
            "analytics_results_loaded user_id=%s row_count=%d student_count=%d",
            request.user.id,
            len(results),
            student_count,
        )

        analysis = PerformanceAnalyzer().analyze(
            AnalysisRequest(
                pdf_id=pdf_id,
                excel_id=excel_id,
                min_students_affected=1,
                confidence_threshold=0.65,
            ),
            classified,
            results,
        )
        store.save_analysis(analysis)

        context["analysis"] = analysis
        context["upload_summary"] = {
            "pdf_id": pdf_id,
            "excel_id": excel_id,
            "questions": len(classified),
            "results": len(results),
            "students": student_count,
        }
        logger.info(
            "analytics_submit_completed user_id=%s weak_topic_count=%d student_summary_count=%d",
            request.user.id,
            len(analysis.class_summary.weak_topics),
            len(analysis.student_summaries),
        )
        messages.success(request, "Class performance analysis completed.")

    except ValueError as exc:
        logger.warning(
            "analytics_submit_validation_failed user_id=%s failure_type=%s",
            request.user.id,
            exc.__class__.__name__,
        )
        messages.error(request, str(exc))
    except RuntimeError as exc:
        error_msg = str(exc)
        logger.error(
            "analytics_submit_runtime_failed user_id=%s failure_type=%s error_msg=%s",
            request.user.id,
            exc.__class__.__name__,
            error_msg[:100],
        )
        if "OPENAI_API_KEY" in error_msg:
            messages.error(
                request,
                "Analytics requires OpenAI API key. Please add OPENAI_API_KEY to Render environment variables."
            )
        else:
            messages.error(request, "Analytics is not fully configured on the backend yet.")
    except Exception as exc:
        logger.error(
            "analytics_submit_failed user_id=%s failure_type=%s status_code=%s",
            request.user.id,
            exc.__class__.__name__,
            getattr(exc, "status_code", "none"),
        )
        messages.error(request, safe_analytics_error_message(exc))

    return render_analytics_response()


@staff_required
def create_folder(request):
    form = FolderForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        folder = form.save(commit=False)
        folder.teacher = request.user
        folder.save()
        messages.success(request, "Folder created.")
        return redirect("teacher_quizzes")
    return render(request, "quizzes/create_folder.html", {"form": form})


@staff_required
def delete_folder(request, folder_id):
    """Delete a subject folder with options to delete or move quizzes"""
    folder = get_object_or_404(SubjectFolder, id=folder_id, teacher=request.user)
    
    if request.method == "POST":
        action = request.POST.get("action")
        
        if action == "delete_all":
            # Delete all quizzes in the folder
            folder.quizzes.all().delete()
            folder.delete()
            messages.success(request, f"Folder '{folder.name}' and all its quizzes have been deleted.")
        elif action == "move_ungrouped":
            # Move quizzes to ungrouped (set folder to None)
            folder.quizzes.all().update(folder=None)
            folder.delete()
            messages.success(request, f"Folder '{folder.name}' deleted. Quizzes moved to ungrouped.")
        else:
            messages.error(request, "Invalid action.")
            return redirect("folder_detail", folder_id=folder_id)
        
        return redirect("teacher_quizzes")
    
    quiz_count = folder.quizzes.count()
    return render(request, "quizzes/delete_folder.html", {
        "folder": folder,
        "quiz_count": quiz_count,
    })


@staff_required
def folder_detail(request, folder_id):
    folder = get_object_or_404(SubjectFolder, id=folder_id, teacher=request.user)
    quizzes = (
        folder.quizzes.all()
        .annotate(
            assigned_count=Count("attempt_permissions", distinct=True),
            submitted_count=Count(
                "submissions",
                filter=Q(submissions__is_submitted=True),
                distinct=True,
            ),
        )
        .order_by("-created_at")
    )
    for quiz in quizzes:
        quiz.bar_max = max(1, quiz.assigned_count, quiz.submitted_count)
    return render(request, "quizzes/folder_detail.html", {
        "folder": folder,
        "quizzes": quizzes,
    })


@staff_required
def folder_analytics(request, folder_id):
    """AI-powered analytics to identify weak topics for students in a folder"""
    folder = get_object_or_404(SubjectFolder, id=folder_id, teacher=request.user)
    
    # Get all quizzes in folder
    quizzes = folder.quizzes.all()
    
    # Collect all wrong answers with question details
    wrong_answers_data = []
    question_stats = {}
    student_performance = {}
    
    for quiz in quizzes:
        submissions = Submission.objects.filter(quiz=quiz, is_submitted=True).select_related('student_user')
        
        for submission in submissions:
            student_id = submission.student_user_id
            if student_id not in student_performance:
                student_name = submission.student_name
                if submission.student_user and hasattr(submission.student_user, 'student_profile'):
                    sp = submission.student_user.student_profile
                    student_name = f"{sp.first_name} {sp.second_name}".strip()
                student_performance[student_id] = {
                    'name': student_name,
                    'total_questions': 0,
                    'correct': 0,
                    'wrong_topics': [],
                }
            
            answers = submission.answers.select_related('question').all()
            for answer in answers:
                q = answer.question
                if q.question_type == 'file_upload':
                    continue  # Skip file upload questions
                
                student_performance[student_id]['total_questions'] += 1
                
                # Track question stats
                if q.id not in question_stats:
                    question_stats[q.id] = {
                        'text': q.text[:100],
                        'quiz_title': quiz.title,
                        'total_attempts': 0,
                        'wrong_count': 0,
                        'question_type': q.question_type,
                    }
                
                question_stats[q.id]['total_attempts'] += 1
                
                if answer.is_correct:
                    student_performance[student_id]['correct'] += 1
                else:
                    question_stats[q.id]['wrong_count'] += 1
                    wrong_answers_data.append({
                        'question': q.text,
                        'quiz': quiz.title,
                        'student': student_performance[student_id]['name'],
                    })
    
    # Calculate difficulty rates
    difficult_questions = []
    for qid, stats in question_stats.items():
        if stats['total_attempts'] > 0:
            error_rate = (stats['wrong_count'] / stats['total_attempts']) * 100
            stats['error_rate'] = round(error_rate, 1)
            if error_rate > 30:  # More than 30% got it wrong
                difficult_questions.append(stats)
    
    # Sort by error rate
    difficult_questions.sort(key=lambda x: x['error_rate'], reverse=True)
    
    # Calculate student performance percentages
    students_list = []
    for sid, data in student_performance.items():
        if data['total_questions'] > 0:
            data['percentage'] = round((data['correct'] / data['total_questions']) * 100, 1)
            students_list.append(data)
    
    # Sort by performance (lowest first = needs most help)
    students_list.sort(key=lambda x: x['percentage'])
    
    # AI Analysis
    ai_analysis = None
    if request.method == 'POST' and 'analyze' in request.POST:
        ai_analysis = _generate_ai_analysis(folder.name, difficult_questions, students_list, wrong_answers_data)
    
    return render(request, "quizzes/folder_analytics.html", {
        "folder": folder,
        "difficult_questions": difficult_questions[:15],  # Top 15 hardest
        "students": students_list,
        "total_submissions": sum(1 for s in students_list),
        "total_questions_analyzed": sum(s['total_attempts'] for s in question_stats.values()),
        "ai_analysis": ai_analysis,
    })


def _generate_ai_analysis(folder_name, difficult_questions, students, wrong_answers_data):
    """Generate AI-powered analysis of weak topics"""
    try:
        import openai
        from django.conf import settings
        
        api_key = getattr(settings, 'OPENAI_API_KEY', None)
        if not api_key:
            return {
                'error': True,
                'message': 'OpenAI API key not configured. Add OPENAI_API_KEY to your settings.'
            }
        
        client = openai.OpenAI(api_key=api_key)
        
        # Prepare context for AI
        difficult_q_text = "\n".join([
            f"- Question: '{q['text']}' (from {q['quiz_title']}) - {q['error_rate']}% error rate"
            for q in difficult_questions[:10]
        ])
        
        struggling_students = [s for s in students if s['percentage'] < 60][:5]
        student_text = "\n".join([
            f"- {s['name']}: {s['percentage']}% correct ({s['correct']}/{s['total_questions']})"
            for s in struggling_students
        ])
        
        prompt = f"""You are an educational analyst. Analyze the following quiz performance data for the subject folder "{folder_name}".

MOST DIFFICULT QUESTIONS (highest error rates):
{difficult_q_text if difficult_q_text else "No significant difficulty patterns found."}

STRUGGLING STUDENTS (below 60%):
{student_text if student_text else "No students below 60% threshold."}

Based on this data, provide:
1. **Key Weak Topics**: Identify 3-5 specific topics/concepts students are struggling with based on the questions they got wrong.
2. **Root Cause Analysis**: What underlying concepts might students be missing?
3. **Recommendations**: Specific teaching strategies or resources to address these gaps.
4. **Priority Actions**: What should the teacher focus on immediately?

Be specific and actionable. Use the actual question content to identify patterns."""

        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {"role": "system", "content": "You are an expert educational analyst who helps teachers identify learning gaps and improve student outcomes."},
                {"role": "user", "content": prompt}
            ],
            max_tokens=1500,
            temperature=0.7
        )
        
        return {
            'error': False,
            'analysis': response.choices[0].message.content
        }
        
    except ImportError:
        return {
            'error': True,
            'message': 'OpenAI library not installed. Run: pip install openai'
        }
    except openai.RateLimitError:
        return {
            'error': True,
            'message': '⚠️ OpenAI API quota exceeded. Please check your OpenAI billing at platform.openai.com or wait until your quota resets.'
        }
    except openai.AuthenticationError:
        return {
            'error': True,
            'message': '⚠️ Invalid OpenAI API key. Please check your OPENAI_API_KEY in settings.'
        }
    except openai.APIConnectionError:
        return {
            'error': True,
            'message': '⚠️ Could not connect to OpenAI. Please check your internet connection.'
        }
    except Exception as e:
        error_msg = str(e)
        if 'insufficient_quota' in error_msg or '429' in error_msg:
            return {
                'error': True,
                'message': '⚠️ OpenAI API quota exceeded. Please add credits at platform.openai.com/account/billing or use a different API key.'
            }
        return {
            'error': True,
            'message': 'AI Analysis failed. Please try again later.'
        }


@staff_required
def quiz_live_counts(request):
    raw_ids = request.GET.get("quiz_ids", "")
    quiz_ids = [int(qid) for qid in raw_ids.split(",") if qid.strip().isdigit()]
    if not quiz_ids:
        return JsonResponse({"quizzes": {}})

    quizzes = (
        Quiz.objects.filter(teacher=request.user, id__in=quiz_ids)
        .annotate(
            assigned_count=Count("attempt_permissions", distinct=True),
            submitted_count=Count(
                "submissions",
                filter=Q(submissions__is_submitted=True),
                distinct=True,
            ),
        )
    )

    payload = {}
    for quiz in quizzes:
        bar_max = max(1, quiz.assigned_count, quiz.submitted_count)
        payload[str(quiz.id)] = {
            "assigned_count": quiz.assigned_count,
            "submitted_count": quiz.submitted_count,
            "bar_max": bar_max,
        }

    return JsonResponse({"quizzes": payload})



@staff_required
def create_quiz(request):
    pre_folder_id = request.GET.get("folder")

    form = QuizForm(request.POST or None, teacher=request.user)

    # Preselect folder when opening the page
    if request.method == "GET" and pre_folder_id:
        form.initial["folder"] = pre_folder_id

    if request.method == "POST" and form.is_valid():
        quiz = form.save(commit=False)
        quiz.teacher = request.user
        quiz.save()
        messages.success(request, f"Quiz created. Code: {quiz.code}")

        # if created inside a folder, go back to that folder
        if quiz.folder_id:
            return redirect("folder_detail", quiz.folder_id)

        return redirect("teacher_quizzes")

    return render(request, "quizzes/create_quiz.html", {"form": form})


@staff_required
def create_question(request, quiz_id):
    quiz = get_object_or_404(Quiz, id=quiz_id, teacher=request.user)
    
    # Get question type from request (default based on quiz type or multiple_choice)
    question_type = request.GET.get('type') or request.POST.get('question_type') or 'multiple_choice'
    
    # Choose the correct form based on selected question type
    if question_type == 'true_false':
        FormClass = TrueFalseQuestionForm
    elif question_type == 'file_upload':
        FormClass = FileUploadQuestionForm
    else:
        FormClass = QuestionForm
    
    form = FormClass(request.POST or None, request.FILES or None)

    if request.method == "POST" and form.is_valid():
        q = form.save(commit=False)
        q.quiz = quiz
        # Set question_type based on selected type
        selected_type = request.POST.get('question_type', 'multiple_choice')
        if selected_type == 'true_false':
            q.question_type = 'true_false'
            q.option1 = 'True'
            q.option2 = 'False'
            q.option3 = ''
            q.option4 = ''
        elif selected_type == 'file_upload':
            q.question_type = 'file_upload'
            q.option1 = ''
            q.option2 = ''
            q.option3 = ''
            q.option4 = ''
            q.correct_option = 1  # Default (not used for file uploads)
        else:
            q.question_type = 'multiple_choice'
        q.save()
        messages.success(request, "Question added.")
        return redirect("teacher_quiz_detail", quiz_id=quiz.id)

    return render(request, "quizzes/create_question.html", {
        "quiz": quiz, 
        "form": form,
        "selected_question_type": question_type,
    })


@staff_required
def quiz_submissions(request, quiz_id):
    quiz = get_object_or_404(Quiz, id=quiz_id, teacher=request.user)

    submissions = list(
        Submission.objects
        .filter(quiz=quiz, is_submitted=True)
        .select_related("student_user", "student_user__student_profile")
        .prefetch_related("security_violations")
        .order_by("-submitted_at", "-id")
    )
    for submission in submissions:
        _attach_security_summary(
            submission,
            prefetched_violations=getattr(submission, "_prefetched_objects_cache", {}).get("security_violations"),
        )

    student_ids = [s.student_user_id for s in submissions if s.student_user_id]

    perms_qs = QuizAttemptPermission.objects.filter(
        quiz=quiz,
        student_user_id__in=student_ids
    ).values("student_user_id", "allowed_attempts")

    attempt_map = {p["student_user_id"]: p["allowed_attempts"] for p in perms_qs}

    return render(request, "quizzes/quiz_submissions.html", {
        "quiz": quiz,
        "submissions": submissions,
        "attempt_map": attempt_map,
    })


@staff_required
def grade_submission(request, quiz_id, submission_id):
    """Teacher view to grade a student's submission with file uploads"""
    quiz = get_object_or_404(Quiz, id=quiz_id, teacher=request.user)
    submission = get_object_or_404(Submission, id=submission_id, quiz=quiz)
    
    # Get answers and file submissions
    answers = submission.answers.select_related('question').order_by('id')
    file_submissions = submission.file_submissions.select_related('question').all()
    violations = list(submission.security_violations.all())
    _attach_security_summary(submission, prefetched_violations=violations)
    
    # Get student info
    student_name = submission.student_name or ""
    university_id = ""
    if submission.student_user_id:
        try:
            sp = getattr(submission.student_user, "student_profile", None)
            if sp:
                student_name = f"{sp.first_name} {sp.second_name} {sp.third_name}".strip() or student_name
                university_id = sp.university_id or ""
        except:
            pass
    
    if request.method == "POST":
        # Handle grading for each file submission (per-question file uploads only)
        for fs in file_submissions:
            grade_key = f"grade_{fs.id}"
            comment_key = f"comment_{fs.id}"
            file_key = f"teacher_file_{fs.id}"
            
            if grade_key in request.POST:
                fs.grade = request.POST.get(grade_key, '').strip() or None
            if comment_key in request.POST:
                fs.teacher_comment = request.POST.get(comment_key, '').strip() or None
            
            # Handle teacher file upload
            if file_key in request.FILES:
                teacher_file = request.FILES[file_key]
                fs.teacher_file = teacher_file
                fs.teacher_file_name = teacher_file.name
            
            fs.graded_at = timezone.now()
            fs.save()
        
        messages.success(request, "Grading saved successfully!")
        return redirect("grade_submission", quiz_id=quiz.id, submission_id=submission.id)
    
    return render(request, "quizzes/grade_submission.html", {
        "quiz": quiz,
        "submission": submission,
        "student_name": student_name,
        "university_id": university_id,
        "answers": answers,
        "file_submissions": file_submissions,
        "security_violations": violations,
    })


@staff_required
@require_POST
def delete_teacher_file(request, file_submission_id):
    """Delete teacher's uploaded feedback file from a FileSubmission"""
    fs = get_object_or_404(FileSubmission, id=file_submission_id)
    quiz = fs.submission.quiz
    
    # Verify teacher owns this quiz
    if quiz.teacher != request.user:
        messages.error(request, "Permission denied.")
        return redirect("teacher_quizzes")
    
    # Delete the file
    if fs.teacher_file:
        fs.teacher_file.delete()
        fs.teacher_file_name = None
        fs.save()
        messages.success(request, "Feedback file removed.")
    
    return redirect("grade_submission", quiz_id=quiz.id, submission_id=fs.submission.id)


@staff_required
@require_POST
def delete_submission_teacher_file(request, submission_id):
    """Delete teacher's uploaded feedback file from a Submission"""
    submission = get_object_or_404(Submission, id=submission_id)
    quiz = submission.quiz
    
    # Verify teacher owns this quiz
    if quiz.teacher != request.user:
        messages.error(request, "Permission denied.")
        return redirect("teacher_quizzes")
    
    # Delete the file
    if submission.teacher_file:
        submission.teacher_file.delete()
        submission.teacher_file_name = None
        submission.save()
        messages.success(request, "Feedback file removed.")
    
    return redirect("grade_submission", quiz_id=quiz.id, submission_id=submission.id)

   
@staff_required
@require_POST
def adjust_attempts(request, quiz_id, student_id):
    quiz = get_object_or_404(Quiz, id=quiz_id, teacher=request.user)
    student = get_object_or_404(User, id=student_id)

    try:
        delta = int(request.POST.get("delta", "0"))
    except ValueError:
        delta = 0

    if delta not in (-1, 1):
        messages.error(request, "Invalid attempt change.")
        return redirect("quiz_submissions", quiz_id=quiz.id)

    perm, _ = QuizAttemptPermission.objects.get_or_create(
        quiz=quiz,
        student_user=student,
        defaults={"allowed_attempts": 1}
    )

    new_val = perm.allowed_attempts + delta
    if new_val < 1:
        new_val = 1  # ✅ never allow 0 or negative

    perm.allowed_attempts = new_val
    perm.save()

    messages.success(request, f"Allowed attempts updated to {new_val} for {student.username}.")
    return redirect("quiz_submissions", quiz_id=quiz.id)
 

@staff_required
def delete_quiz(request, quiz_id):
    quiz = get_object_or_404(Quiz, id=quiz_id, teacher=request.user)

    if request.method == "POST":
        quiz.delete()
        messages.success(request, "Quiz deleted.")
        return redirect("teacher_quizzes")

    return render(request, "quizzes/confirm_delete.html", {"quiz": quiz})


def teacher_logout(request):
    logout(request)
    return redirect('teacher_login')

@transaction.atomic
def student_signup(request):
    if request.user.is_authenticated:
        return redirect("student_dashboard")

    form = StudentSignupForm(request.POST or None)
    next_url = request.GET.get("next") or request.POST.get("next")

    if request.method == "POST":
        if form.is_valid():
            user = form.save()

            StudentProfile.objects.create(
                user=user,
                first_name=form.cleaned_data["first_name"],
                second_name=form.cleaned_data["second_name"],
                third_name=form.cleaned_data["third_name"],
                university_id=form.cleaned_data["university_id"],
                city=form.cleaned_data["city"],
                major=form.cleaned_data["major"],
            )

            login(request, user)
            messages.success(request, "Student account created successfully.")
            if next_url and next_url.startswith("/"):
                return redirect(next_url)
            return redirect("student_dashboard")

        messages.error(request, "Signup failed. Please fix the errors below.")

    return render(request, "quizzes/student_signup.html", {"form": form, "next_url": next_url})


def student_login(request):
    if request.user.is_authenticated:
        return redirect("student_dashboard")

    form = StudentLoginForm(request, data=request.POST or None)
    next_url = request.GET.get("next") or request.POST.get("next")

    if request.method == "POST":
        if form.is_valid():
            user = form.get_user()

            # ✅ Keep teacher accounts out of student login
            if user.is_staff:
                messages.error(request, "This login is for students only.")
                return render(request, "quizzes/student_login.html", {"form": form, "next_url": next_url})

            login(request, user)
            
            # Redirect to next URL if provided and safe, otherwise to dashboard
            if next_url and next_url.startswith('/'):
                return redirect(next_url)
            return redirect("student_dashboard")

        messages.error(request, "Invalid University ID or password.")

    return render(request, "quizzes/student_login.html", {"form": form, "next_url": next_url})


@student_required
def student_dashboard(request):
    profile = request.user.student_profile

    # show history
    submissions = list(
        Submission.objects
        .filter(student_user=request.user)
        .select_related("quiz", "feedback")
        .order_by("-submitted_at")
    )
    for submission in submissions:
        feedback = _get_submission_feedback(submission)
        submission.feedback_ready = _feedback_has_saved_study_set(feedback)
        submission.feedback_flashcard_count = len(feedback.flashcards or []) if feedback else 0
        submission.feedback_practice_count = len(feedback.practice_questions or []) if feedback else 0

    # quiz code form on dashboard
    form = EnterQuizForm(request.POST or None)

    if request.method == "POST" and form.is_valid():
        quiz_code = form.cleaned_data["quiz_code"].strip().upper()
        quiz = Quiz.objects.filter(code=quiz_code).first()
        if not quiz:
            messages.error(request, "Invalid quiz code.")
        else:
            return redirect("take_quiz", quiz_code=quiz.code)

    return render(request, "quizzes/student_dashboard.html", {
        "profile": profile,
        "submissions": submissions,
        "form": form,
    })

from django.contrib.auth.decorators import login_required
from django.shortcuts import render, get_object_or_404
from .models import Submission

@student_required
def student_submission_detail(request, submission_id):
    submission = get_object_or_404(
        Submission.objects.select_related("quiz", "student_user").prefetch_related("answers__question"),
        id=submission_id,
        student_user=request.user,  
    )

    answers = submission.answers.select_related("question").all().order_by("question__id")
    file_submissions = submission.file_submissions.select_related("question").all()
    feedback = _get_submission_feedback(submission)
    feedback_ready = _feedback_has_saved_study_set(feedback)

    return render(request, "quizzes/student_submission_detail.html", {
        "submission": submission,
        "quiz": submission.quiz,
        "answers": answers,
        "file_submissions": file_submissions,
        "feedback": feedback,
        "feedback_ready": feedback_ready,
        "practice_answers_saved": bool(feedback and (feedback.practice_question_attempts or [])),
    })



def app_logout(request):
    logout(request)
    return redirect("home")

@staff_required
def move_quiz(request, quiz_id):
    quiz = get_object_or_404(Quiz, id=quiz_id, teacher=request.user)
    form = MoveQuizForm(request.POST or None, teacher=request.user)

    if request.method == "POST" and form.is_valid():
        quiz.folder = form.cleaned_data["folder"]
        quiz.save()
        messages.success(request, "Quiz moved successfully.")
        return redirect("teacher_quizzes")

    return render(request, "quizzes/move_quiz.html", {
        "quiz": quiz,
        "form": form
    })
def _teacher_required(user):
    return user.is_authenticated and user.is_staff


@login_required
def teacher_quiz_detail(request, quiz_id):
    if not request.user.is_staff:
        return redirect("home")

    quiz = get_object_or_404(Quiz, id=quiz_id, teacher=request.user)
    questions = quiz.questions.order_by("id")
    
    # Calculate submission statistics
    total_submissions = quiz.submissions.count()
    submitted_count = quiz.submissions.filter(is_submitted=True).count()
    not_submitted_count = total_submissions - submitted_count
    
    # Calculate percentages for the graph
    if total_submissions > 0:
        submitted_percentage = (submitted_count / total_submissions) * 100
        not_submitted_percentage = (not_submitted_count / total_submissions) * 100
    else:
        submitted_percentage = 0
        not_submitted_percentage = 0

    return render(
        request,
        "quizzes/teacher_quiz_detail.html",
        {
            "quiz": quiz, 
            "questions": questions,
            "total_submissions": total_submissions,
            "submitted_count": submitted_count,
            "not_submitted_count": not_submitted_count,
            "submitted_percentage": submitted_percentage,
            "not_submitted_percentage": not_submitted_percentage,
        },
    )


@login_required
def edit_question(request, quiz_id, question_id):
    if not request.user.is_staff:
        return redirect("home")

    quiz = get_object_or_404(Quiz, id=quiz_id, teacher=request.user)
    question = get_object_or_404(Question, id=question_id, quiz=quiz)

    form = QuestionForm(request.POST or None, request.FILES or None, instance=question)

    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "Question updated successfully.")
        return redirect("teacher_quiz_detail", quiz_id=quiz.id)

    return render(
        request,
        "quizzes/edit_question.html",
        {"quiz": quiz, "question": question, "form": form},
    )


@login_required
def delete_question(request, quiz_id, question_id):
    if not request.user.is_staff:
        return redirect("home")

    quiz = get_object_or_404(Quiz, id=quiz_id, teacher=request.user)
    question = get_object_or_404(Question, id=question_id, quiz=quiz)

    if request.method == "POST":
        question.delete()
        messages.success(request, "Question deleted.")
        return redirect("teacher_quiz_detail", quiz_id=quiz.id)

    return render(
        request,
        "quizzes/delete_question_confirm.html",
        {"quiz": quiz, "question": question},
    )
    
@staff_required
def edit_quiz_settings(request, quiz_id):
    quiz = get_object_or_404(Quiz, id=quiz_id, teacher=request.user)
    form = QuizSettingsForm(request.POST or None, instance=quiz)

    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "Quiz settings updated.")
        return redirect("teacher_quiz_detail", quiz_id=quiz.id)

    return render(request, "quizzes/edit_quiz_settings.html", {
        "quiz": quiz,
        "form": form
    })

# ---------------------------
# HELPERS
# ---------------------------
def _latest_per_student(subs_queryset):
    """
    Keep latest submission per student for a quiz.
    Key: student_user_id if exists, else student_name.
    """
    latest = {}
    for s in subs_queryset.order_by("-submitted_at", "-id"):
        key = s.student_user_id or (s.student_name or "").strip().lower()
        if key and key not in latest:
            latest[key] = s
    return list(latest.values())


def _student_info(submission):
    """
    Returns: (full_name, university_id)
    """
    full_name = (submission.student_name or "").strip()
    university_id = ""

    u = getattr(submission, "student_user", None)
    sp = getattr(u, "student_profile", None) if u else None

    if sp:
        full_name = " ".join(filter(None, [sp.first_name, sp.second_name, sp.third_name])).strip() or full_name
        university_id = getattr(sp, "university_id", "") or ""

    return full_name, university_id



def _autosize(ws, max_width=45):
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = 0
        for cell in ws[letter]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max_len + 2, max_width)


def _apply_table(ws, start_row, end_row, end_col, table_name):
    ref = f"A{start_row}:{get_column_letter(end_col)}{end_row}"
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def _safe_table_name(name: str) -> str:
    cleaned = "".join(ch if ch.isalnum() else "_" for ch in name)
    if cleaned and cleaned[0].isdigit():
        cleaned = "T_" + cleaned
    return cleaned[:50]



@staff_required
def export_submissions_excel(request, quiz_id):
    quiz = get_object_or_404(Quiz, id=quiz_id, teacher=request.user)

    submissions = (
        Submission.objects
        .filter(quiz=quiz, is_submitted=True)
        .select_related("student_user", "student_user__student_profile")
        .prefetch_related("answers__question", "file_submissions__question")
        .order_by("-submitted_at", "-id")
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Submissions"

      


        


    # ---------------------------
    # Styles (MATCH Student Report)
    # ---------------------------
    dark_title_fill = PatternFill("solid", fgColor="111827")   # very dark
    dark_bar_fill = PatternFill("solid", fgColor="1F2937")     # dark gray
    blue_header_fill = PatternFill("solid", fgColor="2563EB")  # bright blue

    title_font = Font(size=18, bold=True, color="FFFFFF")
    subtitle_font = Font(size=10, color="FFFFFF")
    header_font = Font(bold=True, color="FFFFFF")

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Get quiz questions so the export includes Q1, Q2, Q3... correctness columns.
    questions = list(quiz.questions.all().order_by("id"))

    # Get file upload questions for this quiz
    file_upload_questions = list(quiz.questions.filter(question_type='file_upload').order_by('id'))
    num_file_cols = len(file_upload_questions)
    question_cols = len(questions)
    total_cols = 6 + question_cols + num_file_cols  # Base cols + Q correctness cols + file grade cols
    end_col_letter = get_column_letter(total_cols) if total_cols > 0 else "F"

    # ---------------------------
    # Row 1: Big Title Bar
    # ---------------------------
    ws.merge_cells(f"A1:{end_col_letter}1")
    ws["A1"] = "Quiz Submissions Report"
    ws["A1"].font = title_font
    ws["A1"].alignment = center
    ws["A1"].fill = dark_title_fill
    ws.row_dimensions[1].height = 30

    # ---------------------------
    # Row 2: Subtitle Bar
    # ---------------------------
    ws.merge_cells(f"A2:{end_col_letter}2")
    ws["A2"] = (
        f"Quiz: {quiz.title}  |  Code: {quiz.code}  |  "
        f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M')}  |  Teacher: {request.user.username}"
    )
    ws["A2"].font = subtitle_font
    ws["A2"].alignment = center
    ws["A2"].fill = dark_bar_fill
    ws.row_dimensions[2].height = 18

    # spacer row
    ws.append([])

    # ---------------------------
    # Table Header Row (Blue)
    # ---------------------------
    headers = [
        "Student Full Name",
        "University ID",
        "Score",
        "Total",
        "Percentage",
        "Submitted At",
    ]

    # Add per-question correctness columns: 1 = correct, 0 = wrong/unanswered.
    for i in range(1, len(questions) + 1):
        headers.append(f"Q{i}")
    
    # Add file grade headers for each file upload question
    for i, q in enumerate(file_upload_questions, start=1):
        q_label = f"Q{q.id}" if not q.text else q.text[:15]
        headers.append(f"File Grade ({q_label})")

    ws.append(headers)
    header_row = ws.max_row

    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=c)
        cell.font = header_font
        cell.fill = blue_header_fill
        cell.alignment = center
        cell.border = border

    # ---------------------------
    # Data Rows
    # ---------------------------
    start_data_row = header_row + 1

    for s in submissions:
        full_name, university_id = _student_info(s)
        score = int(s.score or 0)
        total = int(s.total or 0)
        pct = (score / total) if total else 0
        submitted_str = s.submitted_at.strftime("%Y-%m-%d %H:%M") if s.submitted_at else ""

        row_data = [
            full_name,
            university_id,
            score,
            total,
            pct,
            submitted_str,
        ]

        # Add per-question correctness values for every question in order.
        # 1 = correct, 0 = wrong or not answered.
        answer_map = {a.question_id: (1 if a.is_correct else 0) for a in s.answers.all()}
        row_data.extend(answer_map.get(q.id, 0) for q in questions)
        
        # Add file grades for each file upload question
        file_grades_map = {fs.question_id: fs.grade for fs in s.file_submissions.all() if fs.question_id}
        for q in file_upload_questions:
            file_grade = file_grades_map.get(q.id, "")
            row_data.append(file_grade or "Not Graded")
        
        ws.append(row_data)

    end_row = ws.max_row

    # Borders + percentage format
    for r in range(start_data_row, end_row + 1):
        ws.cell(row=r, column=5).number_format = "0.0%"
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).border = border
            ws.cell(row=r, column=c).alignment = left if c in (1,) else center

    # Excel Table style (filters + stripes)
    if end_row >= start_data_row:
        ref = f"A{header_row}:{end_col_letter}{end_row}"
        table = Table(displayName=f"QuizSubs{quiz.id}", ref=ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)

    # Freeze under header
    ws.freeze_panes = f"A{start_data_row}"

    _autosize(ws, max_width=45)


    filename = f"{quiz.code}_submissions_report.xlsx"
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response



# =========================================================
# EXPORT 3: FOLDER BOXES EXPORT (one box per quiz, per-question columns)
# URL: teacher/folders/<folder_id>/export-boxes/
# name="export_folder_boxes_excel"
# =========================================================
@staff_required
def export_folder_boxes_excel(request, folder_id):
    folder = get_object_or_404(SubjectFolder, id=folder_id, teacher=request.user)

    quizzes = (
        Quiz.objects
        .filter(folder=folder, teacher=request.user)
        .order_by("created_at")
    )

    all_subs = (
        Submission.objects
        .filter(quiz__in=quizzes, is_submitted=True)
        .select_related("quiz", "student_user", "student_user__student_profile")
        .prefetch_related("answers__question", "file_submissions__question")
        .order_by("-submitted_at", "-id")
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Folder Boxes"

    # Styles
    title_font = Font(size=16, bold=True, color="FFFFFF")
    title_fill = PatternFill("solid", fgColor="111827")

    section_font = Font(size=12, bold=True, color="FFFFFF")
    section_fill = PatternFill("solid", fgColor="1F2937")

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header
    ws.merge_cells("A1:Z1")
    ws["A1"] = f"Subject Folder Boxes Report — {folder.name}"
    ws["A1"].font = title_font
    ws["A1"].fill = title_fill
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:Z2")
    ws["A2"] = f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M')} | Teacher: {request.user.username}"
    ws["A2"].alignment = center
    ws["A2"].fill = section_fill
    ws["A2"].font = Font(size=10, color="FFFFFF")
    ws.row_dimensions[2].height = 18

    row = 4

    for q in quizzes:
        questions = list(q.questions.all().order_by("id"))
        file_upload_questions = [qq for qq in questions if qq.question_type == 'file_upload']
        q_headers = [f"Q{i}" for i in range(1, len(questions) + 1)]
        
        # Add file grade headers for file upload questions
        file_grade_headers = [f"File Grade (Q{i+1})" for i, qq in enumerate(questions) if qq.question_type == 'file_upload']

        headers = [
            "Student Full Name", "University ID", 
            "Score", "Total", "Percentage"
        ] + q_headers + file_grade_headers

        end_col = len(headers)

        # Box title row
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
        tcell = ws.cell(row=row, column=1, value=f"{q.title}   (Code: {q.code})")
        tcell.font = section_font
        tcell.fill = section_fill
        tcell.alignment = left
        row += 1

        # Header row
        for c, h in enumerate(headers, start=1):
            hc = ws.cell(row=row, column=c, value=h)
            hc.font = header_font
            hc.fill = header_fill
            hc.alignment = center
            hc.border = border

        header_row = row
        row += 1

        # Latest submissions per student
        q_subs = all_subs.filter(quiz=q)
        latest = _latest_per_student(q_subs)
        latest.sort(key=lambda s: (_student_info(s)[0] or "").lower())

        start_data_row = row

        for s in latest:
            full_name, university_id = _student_info(s)
            score = int(s.score or 0)
            total = int(s.total or 0)
            pct = (score / total) if total else 0

            # per-question correctness: 1 correct, 0 wrong/unanswered
            ans_map = {a.question_id: (1 if a.is_correct else 0) for a in s.answers.all()}
            per_q = [ans_map.get(qq.id, 0) for qq in questions]
            
            # file grades for file upload questions
            file_grades_map = {fs.question_id: fs.grade for fs in s.file_submissions.all() if fs.question_id}
            file_grades = [file_grades_map.get(qq.id, "Not Graded") or "Not Graded" for qq in questions if qq.question_type == 'file_upload']

            values = [full_name, university_id, score, total, pct] + per_q + file_grades

            for c, v in enumerate(values, start=1):
                dc = ws.cell(row=row, column=c, value=v)
                dc.border = border
                dc.alignment = center if c >= 4 else left

            ws.cell(row=row, column=5).number_format = "0.00%"

            row += 1

        end_row = row - 1

        # Table styling
        if end_row >= start_data_row:
            ref = f"A{header_row}:{get_column_letter(end_col)}{end_row}"
            table_name = _safe_table_name(f"Quiz_{q.id}_{q.code}")
            table = Table(displayName=table_name, ref=ref)
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            ws.add_table(table)

        row += 2

    ws.freeze_panes = "A4"
    _autosize(ws, max_width=35)

    filename = f"{folder.name}_BOXES_report.xlsx".replace(" ", "_")
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

# ---------------------------
# Autosize helper
# ---------------------------
def _autosize_columns(ws, max_width=45):
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = 0
        for cell in ws[letter]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max_len + 2, max_width)


@staff_required
def export_student_folder_excel(request, folder_id, student_id):
    folder = get_object_or_404(SubjectFolder, id=folder_id, teacher=request.user)
    student = get_object_or_404(User, id=student_id)

    quizzes = (
        Quiz.objects
        .filter(folder=folder, teacher=request.user)
        .order_by("created_at")
    )

    submissions = (
        Submission.objects
        .filter(quiz__in=quizzes, student_user=student, is_submitted=True)
        .select_related("quiz", "student_user", "student_user__student_profile")
        .prefetch_related("file_submissions__question")
        .order_by("quiz__title", "-submitted_at", "-id")
    )

    sp = getattr(student, "student_profile", None)

    wb = Workbook()
    ws = wb.active
    ws.title = "Student Report"

    # ------------------------------------------------------------
    # Column widths (match the clean look)
    # ------------------------------------------------------------
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 22

    # ------------------------------------------------------------
    # Styles (matching your screenshot)
    # ------------------------------------------------------------
    dark_title_fill = PatternFill("solid", fgColor="111827")   # very dark
    dark_bar_fill = PatternFill("solid", fgColor="1F2937")     # dark gray
    blue_header_fill = PatternFill("solid", fgColor="2563EB")  # bright blue

    title_font = Font(size=18, bold=True, color="FFFFFF")
    subtitle_font = Font(size=10, color="FFFFFF")
    section_font = Font(size=12, bold=True, color="FFFFFF")
    header_font = Font(bold=True, color="FFFFFF")

    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ------------------------------------------------------------
    # Row 1: Big Title
    # ------------------------------------------------------------
    ws.merge_cells("A1:F1")
    ws["A1"] = "Student Performance Report"
    ws["A1"].font = title_font
    ws["A1"].alignment = center
    ws["A1"].fill = dark_title_fill
    ws.row_dimensions[1].height = 30

    # ------------------------------------------------------------
    # Row 2: Subtitle bar (Subject Folder + Generated)
    # ------------------------------------------------------------
    ws.merge_cells("A2:F2")
    ws["A2"] = f"Subject Folder: {folder.name}   |   Generated: {timezone.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = subtitle_font
    ws["A2"].alignment = center
    ws["A2"].fill = dark_bar_fill
    ws.row_dimensions[2].height = 18

    # ------------------------------------------------------------
    # Student Info section header
    # ------------------------------------------------------------
    ws.merge_cells("A4:F4")
    ws["A4"] = "Student Info"
    ws["A4"].font = section_font
    ws["A4"].alignment = Alignment(horizontal="left", vertical="center")
    ws["A4"].fill = dark_bar_fill
    ws.row_dimensions[4].height = 18

    bold = Font(bold=True)

    # Student info rows (like screenshot)
    info_rows = [
        ("Student Full Name", f"{sp.first_name} {sp.second_name} {sp.third_name}".strip() if sp else student.username),
        ("University ID", sp.university_id if sp else ""),
        ("City", sp.city if sp else ""),
        ("Major", sp.get_major_display() if sp else ""),
    ]

    r = 5
    for label, value in info_rows:
        ws[f"A{r}"] = label
        ws[f"B{r}"] = value
        ws[f"A{r}"].font = bold

        # borders (only A+B like screenshot)
        for c in range(1, 3):
            ws.cell(row=r, column=c).border = border
            ws.cell(row=r, column=c).alignment = left

        r += 1

    # ------------------------------------------------------------
    # Quiz Results section header
    # ------------------------------------------------------------
    start_table = r + 1
    ws.merge_cells(f"A{start_table}:G{start_table}")
    ws[f"A{start_table}"] = "Quiz Results (This Folder)"
    ws[f"A{start_table}"].font = section_font
    ws[f"A{start_table}"].alignment = Alignment(horizontal="left", vertical="center")
    ws[f"A{start_table}"].fill = dark_bar_fill
    ws.row_dimensions[start_table].height = 18

    # ------------------------------------------------------------
    # Table header row (blue)
    # ------------------------------------------------------------
    headers_row = start_table + 1
    headers = ["Quiz Title", "Quiz Code", "Score", "Total", "Percentage", "File Grades", "Submitted At"]

    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=headers_row, column=i, value=h)
        cell.font = header_font
        cell.fill = blue_header_fill
        cell.alignment = center
        cell.border = border

    # latest attempt per quiz (one row per quiz)
    latest = {}
    for s in submissions:
        if s.quiz_id not in latest:
            latest[s.quiz_id] = s

    ordered = sorted(latest.values(), key=lambda x: (x.quiz.title or "").lower())

    data_row = headers_row + 1

    total_score = 0
    total_total = 0

    for s in ordered:
        score = int(s.score or 0)
        tot = int(s.total or 0)
        pct = (score / tot) if tot else 0
        submitted_str = s.submitted_at.strftime("%Y-%m-%d %H:%M") if s.submitted_at else ""

        total_score += score
        total_total += tot
        
        # Collect file grades for this submission
        file_grades = [fs.grade for fs in s.file_submissions.all() if fs.grade]
        file_grades_str = ", ".join(file_grades) if file_grades else "N/A"

        row_values = [s.quiz.title, s.quiz.code, score, tot, pct, file_grades_str, submitted_str]

        for col_idx, val in enumerate(row_values, start=1):
            cell = ws.cell(row=data_row, column=col_idx, value=val)
            cell.border = border
            cell.alignment = left if col_idx in (1, 7) else center

        ws.cell(row=data_row, column=5).number_format = "0.0%"
        data_row += 1

    last_data_row = data_row - 1

    # Add Excel “Table” styling like screenshot filters
    if last_data_row >= headers_row + 1:
        table_ref = f"A{headers_row}:G{last_data_row}"
        table = Table(displayName="StudentQuizResults", ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)

    # Freeze panes under table header
    ws.freeze_panes = f"A{headers_row+1}"

    # Optional autosize
    _autosize_columns(ws, max_width=45)

    filename = f"{folder.name}_{sp.university_id if sp else student.id}_report.xlsx".replace(" ", "_")
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# =============================================================================
# EXPORT 4: QUIZ PDF (Export quiz as PDF document)
# =============================================================================

@staff_required
def export_quiz_pdf(request, quiz_id):
    """
    Export a quiz (with all its questions) as a professional PDF document.
    Designed for educational systems with professional styling.
    
    The PDF includes:
    - Professional header with institution details
    - Quiz metadata and instructions
    - Numbered questions with clear formatting
    - Options in organized tables
    - Answer key (for teacher reference)
    - Professional footer with page numbers
    """
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, PageBreak, Table, TableStyle, HRFlowable
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.pdfgen import canvas
    from io import BytesIO
    from datetime import datetime
    from html import escape
    import re
    
    quiz = get_object_or_404(Quiz, id=quiz_id, teacher=request.user)
    questions = quiz.questions.all().order_by("id")

    def pdf_text(value, default=""):
        """Escape user-authored content before putting it in ReportLab Paragraphs."""
        if value is None:
            return default
        return escape(str(value), quote=False).replace("\n", "<br/>")

    def plain_text(value, default=""):
        """Return ASCII-safe plain text for ReportLab table cells and filenames."""
        if value is None:
            value = default
        text = str(value)
        return text.encode("ascii", "replace").decode("ascii")

    def safe_filename(value):
        text = plain_text(value, "quiz")
        text = re.sub(r"[^A-Za-z0-9._-]+", "_", text).strip("._")
        return text or "quiz"
    
    # Create PDF in memory with custom page template
    pdf_buffer = BytesIO()
    
    class QuizDocTemplate(SimpleDocTemplate):
        """Custom document template with headers and footers"""
        def __init__(self, *args, **kwargs):
            self.quiz_code = kwargs.pop('quiz_code', '')
            self.quiz_title = kwargs.pop('quiz_title', '')
            super().__init__(*args, **kwargs)
        
        def build(self, flowables, onFirstPage=None, onLaterPages=None, canvasmaker=canvas.Canvas):
            return super().build(
                flowables,
                onFirstPage=self._firstPageHeader,
                onLaterPages=self._laterPageHeader,
                canvasmaker=canvasmaker
            )
        
        def _firstPageHeader(self, canvas, doc):
            self._drawHeader(canvas, doc)
        
        def _laterPageHeader(self, canvas, doc):
            self._drawHeader(canvas, doc)
        
        def _drawHeader(self, canvas, doc):
            canvas.saveState()
            # Header line
            canvas.setLineWidth(2)
            canvas.setStrokeColor(colors.HexColor('#1F4E79'))
            canvas.line(0.5*inch, letter[1]-0.4*inch, letter[0]-0.5*inch, letter[1]-0.4*inch)
            # Footer
            canvas.setFont("Helvetica", 9)
            canvas.setFillColor(colors.HexColor('#666666'))
            canvas.drawString(0.75*inch, 0.4*inch, f"Quiz Code: {self.quiz_code}")
            canvas.drawRightString(letter[0]-0.75*inch, 0.4*inch, f"Page {doc.page}")
            canvas.restoreState()
    
    doc = QuizDocTemplate(
        pdf_buffer,
        pagesize=letter,
        rightMargin=0.75*inch,
        leftMargin=0.75*inch,
        topMargin=1*inch,
        bottomMargin=0.75*inch,
        quiz_code=quiz.code,
        quiz_title=quiz.title
    )
    
    # Define professional styles
    styles = getSampleStyleSheet()
    
    # Institution header
    institution_style = ParagraphStyle(
        'Institution',
        parent=styles['Normal'],
        fontSize=11,
        textColor=colors.HexColor('#1F4E79'),
        alignment=TA_CENTER,
        fontName='Helvetica-Bold',
        spaceAfter=2
    )
    
    # Main title
    title_style = ParagraphStyle(
        'QuizTitle',
        parent=styles['Heading1'],
        fontSize=22,
        textColor=colors.HexColor('#0F3460'),
        spaceAfter=4,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold',
        leading=26
    )
    
    # Metadata header
    meta_style = ParagraphStyle(
        'Metadata',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.HexColor('#555555'),
        spaceAfter=10,
        alignment=TA_CENTER,
        fontName='Helvetica',
        leading=12
    )
    
    # Question number
    question_num_style = ParagraphStyle(
        'QuestionNum',
        parent=styles['Heading2'],
        fontSize=13,
        textColor=colors.HexColor('#FFFFFF'),
        fontName='Helvetica-Bold',
        alignment=TA_CENTER
    )
    
    # Question text
    question_text_style = ParagraphStyle(
        'QuestionText',
        parent=styles['Normal'],
        fontSize=12,
        textColor=colors.HexColor('#1F1F1F'),
        spaceAfter=8,
        fontName='Helvetica-Bold',
        leading=14
    )
    
    # Option text in table
    option_text_style = ParagraphStyle(
        'OptionText',
        parent=styles['Normal'],
        fontSize=11,
        textColor=colors.HexColor('#333333'),
        fontName='Helvetica'
    )
    
    # Answer key style
    answer_style = ParagraphStyle(
        'AnswerKey',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.HexColor('#0F6B2F'),
        fontName='Helvetica-Bold'
    )
    
    # Instruction style
    instruction_style = ParagraphStyle(
        'Instructions',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.HexColor('#555555'),
        spaceAfter=12,
        alignment=TA_LEFT,
        fontName='Helvetica',
        leftIndent=10,
        leading=11
    )
    
    # Build story
    story = []
    
    # === COVER SECTION ===
    story.append(Spacer(1, 0.3*inch))
    story.append(Paragraph("QUIZFY ASSESSMENT SYSTEM", institution_style))
    story.append(Spacer(1, 0.15*inch))
    story.append(Paragraph(pdf_text(quiz.title, "Untitled Quiz"), title_style))
    story.append(Spacer(1, 0.2*inch))
    
    # Metadata table
    quiz_type_display = dict(Quiz.QUIZ_TYPE_CHOICES).get(quiz.quiz_type, quiz.quiz_type)
    metadata = [
        ["Quiz Code:", plain_text(quiz.code)],
        ["Question Type:", plain_text(quiz_type_display)],
        ["Total Questions:", str(questions.count())],
        ["Duration:", f"{quiz.duration_minutes} minutes" if quiz.duration_minutes else "Unlimited"],
        ["Instructor:", plain_text(request.user.get_full_name() or request.user.username)],
        ["Generated:", datetime.now().strftime('%B %d, %Y at %I:%M %p')]
    ]
    
    meta_table = Table(metadata, colWidths=[2.2*inch, 3*inch])
    meta_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#E8F0F7')),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#1F4E79')),
        ('TEXTCOLOR', (1, 0), (1, -1), colors.HexColor('#333333')),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('LEFTPADDING', (0, 0), (-1, -1), 10),
        ('RIGHTPADDING', (0, 0), (-1, -1), 10),
        ('LINEBEFORE', (1, 0), (1, -1), 1, colors.HexColor('#CCCCCC')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E0E0E0')),
        ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.white, colors.HexColor('#F8F8F8')])
    ]))
    story.append(meta_table)
    story.append(Spacer(1, 0.3*inch))
    
    # Instructions section
    story.append(Paragraph("INSTRUCTIONS FOR STUDENTS", institution_style))
    story.append(Spacer(1, 0.1*inch))
    story.append(Paragraph("- Answer all questions carefully and completely", instruction_style))
    story.append(Paragraph("- Select the most appropriate option for each question", instruction_style))
    story.append(Paragraph("- Show all work for written responses", instruction_style))
    story.append(Paragraph("- Time management is important", instruction_style))
    story.append(Spacer(1, 0.15*inch))
    
    story.append(PageBreak())
    
    # === QUESTIONS SECTION ===
    question_count = 0
    items_on_page = 0
    
    for idx, question in enumerate(questions, 1):
        question_count += 1
        
        # Question header with number (colored background)
        q_number_table = Table([["QUESTION " + str(idx)]], colWidths=[doc.width])
        q_number_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#1F4E79')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.white),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 12),
            ('ALIGNMENT', (0, 0), (-1, -1), 'CENTER'),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        story.append(q_number_table)
        story.append(Spacer(1, 0.1*inch))
        
        # Question text
        if question.text:
            story.append(Paragraph(pdf_text(question.text), question_text_style))
        story.append(Spacer(1, 0.08*inch))
        
        # Options based on question type
        if question.question_type == 'multiple_choice':
            options = [question.option1, question.option2, question.option3, question.option4]
            options = [opt for opt in options if opt]
            if options:
                # Create options table
                table_data = []
                for i, option in enumerate(options, 1):
                    letter_opt = chr(64+i)  # A, B, C, D
                    table_data.append([
                        f"{letter_opt}.",
                        Paragraph(pdf_text(option), option_text_style),
                        "Correct" if question.correct_option == i else ""
                    ])
                
                options_table = Table(table_data, colWidths=[0.4*inch, doc.width - 1.5*inch, 1.1*inch])
                options_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#FAFAFA')),
                    ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#333333')),
                    ('TEXTCOLOR', (2, 0), (2, -1), colors.HexColor('#0F6B2F')),
                    ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                    ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('ALIGNMENT', (0, 0), (0, -1), 'CENTER'),
                    ('ALIGNMENT', (2, 0), (2, -1), 'CENTER'),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                    ('LEFTPADDING', (0, 0), (-1, -1), 8),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 8),
                    ('TOPPADDING', (0, 0), (-1, -1), 6),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                    ('LINEBELOW', (0, 0), (-1, -1), 0.5, colors.HexColor('#E0E0E0')),
                ]))
                story.append(options_table)
        
        elif question.question_type == 'true_false':
            # True/False options table
            table_data = [
                ["A.", "True", "Correct" if question.correct_option == 1 else ""],
                ["B.", "False", "Correct" if question.correct_option == 2 else ""]
            ]
            tf_table = Table(table_data, colWidths=[0.4*inch, doc.width - 1.5*inch, 1.1*inch])
            tf_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#FAFAFA')),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#333333')),
                ('TEXTCOLOR', (2, 0), (2, -1), colors.HexColor('#0F6B2F')),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('ALIGNMENT', (0, 0), (0, -1), 'CENTER'),
                ('ALIGNMENT', (2, 0), (2, -1), 'CENTER'),
                ('LEFTPADDING', (0, 0), (-1, -1), 8),
                ('RIGHTPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('LINEBELOW', (0, 0), (-1, -1), 0.5, colors.HexColor('#E0E0E0')),
            ]))
            story.append(tf_table)
        
        elif question.question_type == 'file_upload':
            upload_table = Table([["File Upload Question - Students submit PDF, JPG, or PNG"]], colWidths=[doc.width])
            upload_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#FFF4E6')),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#E67E22')),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Oblique'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('ALIGNMENT', (0, 0), (-1, -1), 'CENTER'),
                ('LEFTPADDING', (0, 0), (-1, -1), 6),
                ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('BORDER', (0, 0), (-1, -1), 1, colors.HexColor('#E67E22'))
            ]))
            story.append(upload_table)
        
        story.append(Spacer(1, 0.25*inch))
        items_on_page += 1
        
        # Page break after 2-3 questions for readability
        if (idx % 3 == 0) and (idx < questions.count()):
            story.append(PageBreak())
            items_on_page = 0
    
    # === ANSWER KEY (LAST PAGE) ===
    story.append(PageBreak())
    story.append(Paragraph("ANSWER KEY (TEACHER REFERENCE)", institution_style))
    story.append(Spacer(1, 0.2*inch))
    
    # Build answer key table
    answer_data = [["Question", "Type", "Correct Answer"]]
    for idx, question in enumerate(questions, 1):
        if question.question_type == 'multiple_choice':
            correct = chr(64 + question.correct_option)
            answer = f"{correct}) {plain_text(getattr(question, f'option{question.correct_option}', 'N/A'))}"
        elif question.question_type == 'true_false':
            answer = "True" if question.correct_option == 1 else "False"
        else:
            answer = "Manual Review"
        
        answer_data.append([
            str(idx),
            plain_text(question.get_question_type_display()),
            plain_text(answer)
        ])
    
    answer_table = Table(answer_data, colWidths=[0.8*inch, 2*inch, doc.width - 2.8*inch])
    answer_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E79')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor('#333333')),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('ALIGNMENT', (0, 0), (0, -1), 'CENTER'),
        ('ALIGNMENT', (1, 0), (1, -1), 'CENTER'),
        ('ALIGNMENT', (2, 0), (2, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#CCCCCC')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F8F8')])
    ]))
    story.append(answer_table)
    story.append(Spacer(1, 0.3*inch))
    story.append(Paragraph("This answer key is for instructor reference only.", meta_style))
    
    # Build PDF
    doc.build(story)
    
    # Return as downloadable file
    pdf_buffer.seek(0)
    response = HttpResponse(pdf_buffer.getvalue(), content_type='application/pdf')
    filename = f"{safe_filename(quiz.code)}_{safe_filename(quiz.title)}_Assessment.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


def quiz_status(request, quiz_code):
    quiz = get_object_or_404(Quiz, code=quiz_code)
    return JsonResponse({"active": bool(quiz.is_active)})
@staff_member_required
def send_test_email(request):
    to_email = request.user.email
    if not to_email:
        messages.error(request, "Your teacher account has no email saved in User.email.")
        return redirect("teacher_quizzes")

    try:
        send_mail(
            subject="Quizfy SMTP Test ✅",
            message="If you received this, Render + Gmail SMTP works.",
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=[to_email],
            fail_silently=False,
        )
        messages.success(request, f"Test email sent to {to_email}. Check inbox/spam.")
    except Exception as e:
        messages.error(request, f"SMTP failed: {e}")
    return redirect("teacher_quizzes")
def test_send_email(request):
    send_mail(
        subject="Quizfy Test Email",
        message="Hello! This is a test from Quizfy via SendGrid.",
        from_email=settings.DEFAULT_FROM_EMAIL,
        recipient_list=["YOUR_PERSONAL_EMAIL@gmail.com"],
        fail_silently=False,
    )
    return HttpResponse("Sent (check logs + inbox)")


# ============================================================
# PASSWORD CHANGE FOR LOGGED-IN USERS (Email Verification)
# ============================================================

@login_required
def change_password(request):
    """
    View for logged-in users to change their password.
    Requires old password verification.
    """
    form = ChangePasswordForm(request.user, request.POST or None)
    
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "Your password has been changed successfully!")
        return redirect("change_password_done")
    
    return render(request, "quizzes/change_password.html", {"form": form})


@login_required
def change_password_done(request):
    """
    Confirmation page after password change.
    """
    return render(request, "quizzes/change_password_done.html")


# DIAGNOSTIC ENDPOINT - For troubleshooting email
def email_diagnostic(request):
    """
    Display email configuration for debugging.
    Only accessible if DEBUG=True (safe on production).
    """
    import os
    from django.contrib.sites.models import Site
    
    if not settings.DEBUG:
        return JsonResponse({"error": "Not available in production"}, status=403)
    
    pwd = settings.EMAIL_HOST_PASSWORD
    
    config = {
        "EMAIL_BACKEND": settings.EMAIL_BACKEND,
        "EMAIL_HOST": settings.EMAIL_HOST,
        "EMAIL_PORT": settings.EMAIL_PORT,
        "EMAIL_USE_TLS": settings.EMAIL_USE_TLS,
        "EMAIL_HOST_USER": settings.EMAIL_HOST_USER,
        "EMAIL_HOST_PASSWORD_SET": bool(pwd),
        "EMAIL_HOST_PASSWORD_LENGTH": len(pwd) if pwd else 0,
        "DEFAULT_FROM_EMAIL": settings.DEFAULT_FROM_EMAIL,
        "SITE_ID": getattr(settings, 'SITE_ID', None),
    }
    
    # Get site info
    try:
        sites = list(Site.objects.values('id', 'domain', 'name'))
        config["SITES"] = sites
    except:
        config["SITES"] = "Error fetching sites"
    
    # Try to send test email
    try:
        from django.core.mail import send_mail
        result = send_mail(
            subject="Quizfy Email Test",
            message="Test email from Quizfy diagnostics",
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=["test@example.com"],
            fail_silently=False,
        )
        config["TEST_EMAIL_RESULT"] = f"Success (returned {result})"
    except Exception as e:
        config["TEST_EMAIL_RESULT"] = f"Error: {type(e).__name__}: {str(e)}"
    
    return JsonResponse(config, indent=2)
